"""Excel writer for B2B Dispatch backend audit workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from inventory_cover.b2b_dispatch_schemas import (
    B2B_DUPLICATE_HEADERS,
    B2B_MASTER_HEADERS,
    B2B_RUN_SUMMARY_HEADERS,
    B2B_SHEET_AUDIT_HEADERS,
    B2B_VALIDATION_ISSUE_HEADERS,
    B2BDuplicateRecord,
    B2BSheetAuditRecord,
    B2BValidationIssue,
    NormalizedB2BDispatchRow,
)
from inventory_cover.exceptions import OutputWriteError


def write_b2b_dispatch_report(
    output_path: Path,
    rows: list[NormalizedB2BDispatchRow],
    run_summary: dict[str, Any],
    sheet_audit: list[B2BSheetAuditRecord],
    validation_issues: list[B2BValidationIssue],
    duplicates: list[B2BDuplicateRecord],
) -> None:
    """Write the B2B backend audit workbook using an atomic temp rename."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        master_ws = wb.active
        master_ws.title = "B2B_Dispatch_Master"
        _write_sheet(
            master_ws,
            B2B_MASTER_HEADERS,
            [row.as_master_row() for row in rows],
            "B2BDispatchMasterTable",
        )
        _format_b2b_master_sheet(master_ws)

        summary_ws = wb.create_sheet("Run_Summary")
        _write_sheet(
            summary_ws,
            B2B_RUN_SUMMARY_HEADERS,
            [[run_summary.get(header) for header in B2B_RUN_SUMMARY_HEADERS]],
            "B2BDispatchRunSummaryTable",
        )
        _format_summary_sheet(summary_ws)

        audit_ws = wb.create_sheet("Sheet_Audit")
        _write_sheet(
            audit_ws,
            B2B_SHEET_AUDIT_HEADERS,
            [record.as_row() for record in sheet_audit],
            "B2BDispatchSheetAuditTable",
        )

        issues_ws = wb.create_sheet("Validation_Issues")
        _write_sheet(
            issues_ws,
            B2B_VALIDATION_ISSUE_HEADERS,
            [issue.as_row() for issue in validation_issues],
            "B2BDispatchValidationIssuesTable",
        )

        duplicate_ws = wb.create_sheet("Duplicates")
        _write_sheet(
            duplicate_ws,
            B2B_DUPLICATE_HEADERS,
            [record.as_row() for record in duplicates],
            "B2BDispatchDuplicatesTable",
        )

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write output workbook {output_path}: {exc}") from exc


def _write_sheet(ws: Any, headers: Iterable[str], rows: Iterable[Iterable[Any]], table_name: str) -> None:
    headers = list(headers)
    _assert_unique_headers(headers, ws.title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    _style_sheet(ws, table_name)


def _style_sheet(ws: Any, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    last_row = max(ws.max_row, 1)
    last_col = max(ws.max_column, 1)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = table_ref
    _autosize_columns(ws)


def _format_b2b_master_sheet(ws: Any) -> None:
    header_positions = {cell.value: cell.column for cell in ws[1]}
    for header in (
        "Run ID",
        "Source File",
        "Source Sheet",
        "Source Channel",
        "Appointment ID",
        "Invoice No",
        "PO",
        "Ship To Location",
        "ASIN",
        "PO ASIN Key",
        "Model Number",
        "Dispatch Location",
    ):
        _format_column(ws, header_positions, header, "@")
    for header in ("PO Date", "Dispatch Date"):
        _format_column(ws, header_positions, header, "dd-mmm-yyyy")
    for header in ("Boxes", "PO Qty", "Dispatch Qty"):
        _format_column(ws, header_positions, header, "0")
    for header in (
        "Unit Value",
        "Dispatch Value Source",
        "Dispatch Value Derived",
        "Dispatch Value Difference",
    ):
        _format_column(ws, header_positions, header, "#,##0.00")
    for header in ("Row Validation Notes", "Model Number"):
        col = header_positions.get(header)
        if col is None:
            continue
        ws.column_dimensions[get_column_letter(col)].width = 36
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].alignment = Alignment(wrap_text=True, vertical="top")


def _format_summary_sheet(ws: Any) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        ws.column_dimensions[letter].width = min(max(len(str(column[0].value)) + 2, 14), 36)


def _format_column(ws: Any, header_positions: dict[str, int], header: str, number_format: str) -> None:
    col = header_positions.get(header)
    if col is None:
        return
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        row[0].number_format = number_format


def _autosize_columns(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 55)


def _assert_unique_headers(headers: list[str], sheet_name: str) -> None:
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise OutputWriteError(f"Duplicate output headers in {sheet_name}: {duplicates}")
