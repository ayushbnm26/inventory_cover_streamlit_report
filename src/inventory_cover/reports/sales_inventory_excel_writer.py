"""Excel writers for Vendor Central Sales & Inventory backend artifacts."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from inventory_cover.exceptions import OutputWriteError
from inventory_cover.sales_inventory_schemas import (
    DUPLICATE_HEADERS,
    FILE_AUDIT_HEADERS,
    INVENTORY_MASTER_HEADERS,
    INVENTORY_REPORT_TYPE,
    MAPPING_AUDIT_HEADERS,
    PROCESSING_GUIDE_HEADERS,
    RUN_SUMMARY_HEADERS,
    SALES_MASTER_HEADERS,
    SALES_REPORT_TYPE,
    SOURCE_STATUS_HEADERS,
    VALIDATION_ISSUE_HEADERS,
    NormalizedSalesInventoryRow,
    ReportType,
    SalesInventoryDuplicateRecord,
    SalesInventoryFileAuditRecord,
    SalesInventoryMappingAuditRecord,
    SalesInventoryValidationIssue,
)


def write_sales_backend_report(
    output_path: Path,
    rows: list[NormalizedSalesInventoryRow],
    run_summary: dict[str, Any],
    file_audit: list[SalesInventoryFileAuditRecord],
    validation_issues: list[SalesInventoryValidationIssue],
    duplicates: list[SalesInventoryDuplicateRecord],
    mapping_audit: list[SalesInventoryMappingAuditRecord] | None = None,
) -> None:
    """Write the sales backend audit workbook using an atomic temp rename."""

    _write_backend_report(
        output_path=output_path,
        report_type=SALES_REPORT_TYPE,
        master_sheet_name="Sales_Master",
        master_headers=SALES_MASTER_HEADERS,
        master_table_name="SalesMasterTable",
        run_summary_table_name="SalesRunSummaryTable",
        file_audit_table_name="SalesFileAuditTable",
        validation_table_name="SalesValidationIssuesTable",
        duplicates_table_name="SalesDuplicatesTable",
        mapping_table_name="SalesMappingAuditTable",
        guide_table_name="SalesProcessingGuideTable",
        rows=rows,
        run_summary=run_summary,
        file_audit=file_audit,
        validation_issues=validation_issues,
        duplicates=duplicates,
        mapping_audit=mapping_audit or [],
    )


def write_inventory_backend_report(
    output_path: Path,
    rows: list[NormalizedSalesInventoryRow],
    run_summary: dict[str, Any],
    file_audit: list[SalesInventoryFileAuditRecord],
    validation_issues: list[SalesInventoryValidationIssue],
    duplicates: list[SalesInventoryDuplicateRecord],
    mapping_audit: list[SalesInventoryMappingAuditRecord] | None = None,
) -> None:
    """Write the inventory backend audit workbook using an atomic temp rename."""

    _write_backend_report(
        output_path=output_path,
        report_type=INVENTORY_REPORT_TYPE,
        master_sheet_name="Inventory_Master",
        master_headers=INVENTORY_MASTER_HEADERS,
        master_table_name="InventoryMasterTable",
        run_summary_table_name="InventoryRunSummaryTable",
        file_audit_table_name="InventoryFileAuditTable",
        validation_table_name="InventoryValidationIssuesTable",
        duplicates_table_name="InventoryDuplicatesTable",
        mapping_table_name="InventoryMappingAuditTable",
        guide_table_name="InventoryProcessingGuideTable",
        rows=rows,
        run_summary=run_summary,
        file_audit=file_audit,
        validation_issues=validation_issues,
        duplicates=duplicates,
        mapping_audit=mapping_audit or [],
    )


def write_sales_inventory_run_summary(
    output_path: Path,
    run_summary: dict[str, Any],
    source_status_rows: list[dict[str, Any]],
    validation_issues: list[SalesInventoryValidationIssue],
    mapping_audit: list[SalesInventoryMappingAuditRecord] | None = None,
) -> None:
    """Write the combined Pipeline 3 run summary workbook."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        summary_ws = wb.active
        summary_ws.title = "Run_Summary"
        _write_sheet(
            summary_ws,
            RUN_SUMMARY_HEADERS,
            [[run_summary.get(header) for header in RUN_SUMMARY_HEADERS]],
            "SalesInventoryRunSummaryTable",
        )

        status_ws = wb.create_sheet("Source_Status")
        _write_sheet(
            status_ws,
            SOURCE_STATUS_HEADERS,
            [[row.get(header) for header in SOURCE_STATUS_HEADERS] for row in source_status_rows],
            "SalesInventorySourceStatusTable",
        )

        issues_ws = wb.create_sheet("Validation_Issues")
        _write_sheet(
            issues_ws,
            VALIDATION_ISSUE_HEADERS,
            [issue.as_row() for issue in validation_issues],
            "SalesInventoryValidationIssuesTable",
        )

        mapping_ws = wb.create_sheet("Mapping_Audit")
        _write_sheet(
            mapping_ws,
            MAPPING_AUDIT_HEADERS,
            [record.as_row() for record in (mapping_audit or [])],
            "SalesInventoryMappingAuditTable",
        )

        guide_ws = wb.create_sheet("Processing_Guide")
        _write_sheet(
            guide_ws,
            PROCESSING_GUIDE_HEADERS,
            _processing_guide_rows("COMBINED", [], run_summary),
            "SalesInventoryProcessingGuideTable",
        )
        _format_guide_sheet(guide_ws)

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write output workbook {output_path}: {exc}") from exc


def _write_backend_report(
    output_path: Path,
    report_type: ReportType,
    master_sheet_name: str,
    master_headers: tuple[str, ...],
    master_table_name: str,
    run_summary_table_name: str,
    file_audit_table_name: str,
    validation_table_name: str,
    duplicates_table_name: str,
    mapping_table_name: str,
    guide_table_name: str,
    rows: list[NormalizedSalesInventoryRow],
    run_summary: dict[str, Any],
    file_audit: list[SalesInventoryFileAuditRecord],
    validation_issues: list[SalesInventoryValidationIssue],
    duplicates: list[SalesInventoryDuplicateRecord],
    mapping_audit: list[SalesInventoryMappingAuditRecord],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        master_ws = wb.active
        master_ws.title = master_sheet_name
        _write_sheet(master_ws, master_headers, [row.as_master_row() for row in rows], master_table_name)
        _format_master_sheet(master_ws)

        summary_ws = wb.create_sheet("Run_Summary")
        _write_sheet(
            summary_ws,
            RUN_SUMMARY_HEADERS,
            [[run_summary.get(header) for header in RUN_SUMMARY_HEADERS]],
            run_summary_table_name,
        )

        audit_ws = wb.create_sheet("File_Audit")
        _write_sheet(
            audit_ws,
            FILE_AUDIT_HEADERS,
            [record.as_row() for record in file_audit],
            file_audit_table_name,
        )

        issues_ws = wb.create_sheet("Validation_Issues")
        _write_sheet(
            issues_ws,
            VALIDATION_ISSUE_HEADERS,
            [issue.as_row() for issue in validation_issues],
            validation_table_name,
        )
        _format_long_text_columns(issues_ws, ("Issue Detail", "Action Taken", "Raw Value"))

        duplicate_ws = wb.create_sheet("Duplicates")
        _write_sheet(
            duplicate_ws,
            DUPLICATE_HEADERS,
            [record.as_row() for record in duplicates],
            duplicates_table_name,
        )
        _format_long_text_columns(duplicate_ws, ("Duplicate Key",))

        mapping_ws = wb.create_sheet("Mapping_Audit")
        _write_sheet(
            mapping_ws,
            MAPPING_AUDIT_HEADERS,
            [record.as_row() for record in mapping_audit],
            mapping_table_name,
        )

        guide_ws = wb.create_sheet("Processing_Guide")
        _write_sheet(
            guide_ws,
            PROCESSING_GUIDE_HEADERS,
            _processing_guide_rows(report_type, file_audit, run_summary),
            guide_table_name,
        )
        _format_guide_sheet(guide_ws)

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write output workbook {output_path}: {exc}") from exc


def _write_sheet(ws: Any, headers: Iterable[str], rows: Iterable[Iterable[Any]], table_name: str) -> None:
    headers = list(headers)
    _assert_unique_headers(headers, ws.title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    _style_sheet(ws, table_name)


def _style_sheet(ws: Any, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    last_row = max(ws.max_row, 1)
    last_col = max(ws.max_column, 1)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = table_ref
    _autosize_columns(ws)


def _format_master_sheet(ws: Any) -> None:
    header_positions = {cell.value: cell.column for cell in ws[1]}
    for header in (
        "Run ID",
        "Source File",
        "Source Sheet",
        "Report Type",
        "Programme",
        "Distributor View",
        "View By",
        "Country",
        "Currency",
        "ASIN",
        "Child Vendor Code",
        "Product Title",
        "Brand Code",
        "Brand",
        "Category",
        "Subcategory",
        "Parent ASIN",
        "UPC",
        "EAN",
        "ISBN",
        "Model Number",
        "Store Code",
        "Binding",
        "Colour",
        "Replenishment Code",
    ):
        _format_column(ws, header_positions, header, "@")
    for header in ("Viewing Range Start", "Viewing Range End", "Report Updated Date", "Release Date"):
        _format_column(ws, header_positions, header, "dd-mmm-yyyy")
    for header in (
        "MSRP",
        "Shipped Revenue",
        "Shipped COGS",
        "Sales Discount",
        "Contra-COGS",
        "Net Received",
        "Aged 90+ Days Sellable Inventory",
        "Sellable On-Hand Inventory",
        "Unsellable On-Hand Inventory",
    ):
        _format_column(ws, header_positions, header, "#,##0.00")
    for header in (
        "Shipped Units",
        "Customer Returns",
        "Confirmed Units",
        "Net Received Units",
        "Open Purchase Order Quantity",
        "Overall Vendor Lead Time (days)",
        "Aged 90+ Days Sellable Units",
        "Sellable On Hand Units",
        "Unsellable On-Hand Units",
        "In Transit Quantity",
        "Sellable In Transit Units",
        "Unsellable In Transit Units",
    ):
        _format_column(ws, header_positions, header, "#,##0")
    for header in ("Net PPM %", "ASIN Confirmation %", "Vendor Confirmation %", "Receive Fill %"):
        _format_column(ws, header_positions, header, "0.00")
    _format_long_text_columns(ws, ("Product Title", "Row Validation Notes"))


def _format_column(ws: Any, header_positions: dict[str, int], header: str, number_format: str) -> None:
    col = header_positions.get(header)
    if col is None:
        return
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        row[0].number_format = number_format


def _format_long_text_columns(ws: Any, headers: tuple[str, ...]) -> None:
    header_positions = {cell.value: cell.column for cell in ws[1]}
    for header in headers:
        col = header_positions.get(header)
        if col is None:
            continue
        ws.column_dimensions[get_column_letter(col)].width = 42
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].alignment = Alignment(wrap_text=True, vertical="top")


def _format_guide_sheet(ws: Any) -> None:
    _format_long_text_columns(ws, ("Explanation", "Operational Meaning"))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 64


def _autosize_columns(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 55)


def _processing_guide_rows(
    report_type: str,
    file_audit: list[SalesInventoryFileAuditRecord],
    run_summary: dict[str, Any],
) -> list[list[str]]:
    source_files = ", ".join(record.source_file for record in file_audit) or "No source file for this workbook."
    copied_paths = ", ".join(record.copied_run_path for record in file_audit if record.copied_run_path) or "Not applicable."
    header_rows = ", ".join(
        f"{record.source_file}: row {record.header_row_found}" for record in file_audit if record.header_row_found
    ) or "No header row was detected."
    missing_columns = "; ".join(
        f"{record.source_file}: {record.missing_expected_columns or 'none'}" for record in file_audit
    ) or "None."
    extra_columns = "; ".join(
        f"{record.source_file}: {record.extra_source_columns or 'none'}" for record in file_audit
    ) or "None."
    return [
        [
            "Workbook Purpose",
            "Backend audit artifact",
            f"This workbook stores normalized Vendor Central {report_type.lower()} data for run {run_summary.get('Run ID')}.",
            "It gives the future calculation engine a stable, traceable source table.",
        ],
        [
            "Pipeline Scope",
            "Source ingestion only",
            "The pipeline reads raw Amazon Vendor Central sales and inventory exports, standardizes fields, validates gently, and writes backend artifacts.",
            "It does not calculate inventory-cover metrics.",
        ],
        [
            "Input Folder",
            "Configured source location",
            f"Sales input folder: {run_summary.get('Input sales folder')}; inventory input folder: {run_summary.get('Input inventory folder')}.",
            "Sales and inventory files stay separate from PO Items, B2B Dispatch, outputs, logs, and archives.",
        ],
        [
            "Raw Input Copy",
            "Run-folder evidence",
            f"Raw report(s) processed: {source_files}. Copied input path(s): {copied_paths}.",
            "Original source files are not mutated; run copies support later audit.",
        ],
        [
            "Header Detection",
            "First 20 rows scanned",
            f"Detected header row(s): {header_rows}. Headers are matched after normalizing case, spacing, punctuation, hyphens, underscores, and percent signs.",
            "Small changes in Amazon export formatting should not break the run.",
        ],
        [
            "Metadata Parsing",
            "Report filters",
            "Programme, distributor view, country, currency, viewing range, and report-updated date are parsed from the metadata row when available.",
            "Metadata parsing warnings do not stop the run; source text remains available through audit artifacts.",
        ],
        [
            "Column Standardization",
            "Canonical schema",
            "Expected Vendor Central columns are written into a stable master schema. Missing optional columns are left blank and reported as warnings.",
            "Downstream code can rely on consistent output headers.",
        ],
        [
            "Text Preservation",
            "Identifiers remain text",
            "ASIN, UPC, EAN, ISBN, Model Number, Child Vendor Code, and related identifiers are preserved as text.",
            "Leading zeroes are protected when the source workbook stores them as text or fixed-width values.",
        ],
        [
            "Mapping Enrichment",
            "Blank identifiers only",
            "When a mapping workbook is available, blank Model Number values may be filled from ASIN, and blank ASIN values may be filled from Model Number or SKU. Existing source identifiers are never overwritten.",
            "Enrichment is deterministic: ambiguous mapping keys are skipped and audited instead of guessed.",
        ],
        [
            "Date Parsing",
            "Day-first policy",
            "Release Date, Viewing Range Start, Viewing Range End, and Report Updated Date accept Excel dates and common day-first text formats.",
            "Invalid release dates are warned and kept as source text instead of rejecting the row.",
        ],
        [
            "Number Parsing",
            "No invented zeroes",
            "Numeric strings with commas are parsed safely. Blank numeric cells are written as blank, not zero.",
            "Business logic can later decide how blanks should affect calculations.",
        ],
        [
            "Missing Columns",
            "Allowed with warnings",
            f"Missing expected column summary: {missing_columns}.",
            "A file fails only when too few identifying headers are present to recognize the report type.",
        ],
        [
            "Extra Columns",
            "Ignored but logged",
            f"Extra source column summary: {extra_columns}.",
            "Unexpected Amazon columns do not break ingestion.",
        ],
        [
            "Row Acceptance",
            "Gentle identifier policy",
            "Rows with ASIN are accepted. Rows without ASIN but with Model Number and Product Title are accepted with warning.",
            "The pipeline preserves usable rows for the future inventory-cover engine.",
        ],
        [
            "Rejected Rows",
            "Unusable rows only",
            "Rows are rejected when ASIN, Model Number, and Product Title are all blank.",
            "Every rejected row is listed in Validation_Issues with the action taken.",
        ],
        [
            "Duplicate Handling",
            "Audited by default",
            "Duplicate keys are recorded in Duplicates. Rows are kept by default; exact normalized repeats are dropped only when dedupe is enabled.",
            "No duplicate is silently removed.",
        ],
        [
            "Validation Issues",
            "Warnings over hard failures",
            "Missing optional columns, invalid release dates, invalid numbers, negative quantities, missing ASIN on otherwise usable rows, and duplicates are warnings.",
            "Critical structural failures stop the run; data quality concerns remain visible without over-policing the export.",
        ],
        [
            "Backend-only Nature",
            "Not team-facing",
            "This workbook is a backend artifact for automation and audit. It is not formatted as a team-facing report.",
            "Business users should not treat this workbook as the final inventory-cover report.",
        ],
        [
            "Use by Future Calculation Engine",
            "Source foundation",
            "The upcoming inventory-cover engine will consume latest Pipeline 1, Pipeline 2, and Pipeline 3 backend outputs.",
            "This pipeline completes the sales and inventory source-ingestion foundation.",
        ],
        [
            "What This Workbook Does Not Do",
            "No cover calculations",
            "This pipeline does not calculate DRR, DOH, DOC, stock gap, cover days, recommended PO quantity, or alert flags.",
            "Those calculations belong in the next core inventory-cover calculation engine.",
        ],
    ]


def _assert_unique_headers(headers: list[str], sheet_name: str) -> None:
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise OutputWriteError(f"Duplicate output headers in {sheet_name}: {duplicates}")
