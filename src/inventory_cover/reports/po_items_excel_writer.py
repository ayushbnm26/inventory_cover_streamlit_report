"""Excel writer for the consolidated PO Items report."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from inventory_cover.exceptions import OutputWriteError
from inventory_cover.schemas import (
    DUPLICATE_HEADERS,
    FILE_AUDIT_HEADERS,
    MASTER_HEADERS,
    RUN_SUMMARY_HEADERS,
    TEAM_WORKBOOK_HEADERS,
    VALIDATION_ISSUE_HEADERS,
    DuplicateRecord,
    FileAuditRecord,
    NormalizedPoItemRow,
    ValidationIssue,
)


def write_po_items_report(
    output_path: Path,
    rows: list[NormalizedPoItemRow],
    run_summary: dict[str, Any],
    file_audit: list[FileAuditRecord],
    validation_issues: list[ValidationIssue],
    duplicates: list[DuplicateRecord],
) -> None:
    """Write the final consolidated workbook using an atomic temp rename."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        master_ws = wb.active
        master_ws.title = "PO_Items_Master"
        _write_sheet(master_ws, MASTER_HEADERS, [row.as_master_row() for row in rows], "POItemsMasterTable")
        _format_master_sheet(master_ws)

        summary_ws = wb.create_sheet("Run_Summary")
        summary_rows = [[key, value] for key, value in run_summary.items()]
        _write_sheet(summary_ws, RUN_SUMMARY_HEADERS, summary_rows, "RunSummaryTable")

        audit_ws = wb.create_sheet("File_Audit")
        _write_sheet(audit_ws, FILE_AUDIT_HEADERS, [record.as_row() for record in file_audit], "FileAuditTable")

        issues_ws = wb.create_sheet("Validation_Issues")
        _write_sheet(
            issues_ws,
            VALIDATION_ISSUE_HEADERS,
            [issue.as_row() for issue in validation_issues],
            "ValidationIssuesTable",
        )

        duplicate_ws = wb.create_sheet("Duplicates")
        _write_sheet(
            duplicate_ws,
            DUPLICATE_HEADERS,
            [record.as_row() for record in duplicates],
            "DuplicatesTable",
        )

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write output workbook {output_path}: {exc}") from exc


def write_po_items_team_workbook(
    output_path: Path,
    rows: list[NormalizedPoItemRow],
) -> None:
    """Write the simplified team-facing PO Items workbook."""

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        po_ws = wb.active
        po_ws.title = "PO_Items"
        _write_sheet(po_ws, TEAM_WORKBOOK_HEADERS, [_team_row(row) for row in rows], "POItemsTeamTable")
        _format_master_sheet(po_ws)

        guide_ws = wb.create_sheet("Column_Guide")
        _write_sheet(guide_ws, TEAM_GUIDE_HEADERS, TEAM_GUIDE_ROWS, "ColumnGuideTable")
        _format_guide_sheet(guide_ws)

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write team workbook {output_path}: {exc}") from exc


TEAM_GUIDE_HEADERS: tuple[str, ...] = ("Section", "Column", "Explanation", "How It Is Populated")

TEAM_GUIDE_ROWS: tuple[tuple[str, str, str, str], ...] = (
    (
        "Workbook Purpose",
        "",
        "This workbook combines Amazon PO Items files into one clean list for business use.",
        "Each accepted row from the input files is copied into the PO_Items sheet after standard formatting and checks.",
    ),
    (
        "Row Normalization",
        "",
        "Rows are standardized so that the same business fields appear in the same columns, even when Amazon export headers vary slightly.",
        "The process reads recognized PO Item headers, keeps all valid rows, standardizes dates and numbers, and leaves a blank Remarks column for team notes.",
    ),
    (
        "Date Normalization",
        "Window Start / Window End / Expected Date",
        "Dates are shown in a consistent date format.",
        "Window Start and Window End are read as day/month/year. Expected Date is read as month/day/year, matching the Amazon export behavior.",
    ),
    (
        "Number Normalization",
        "Quantity and Cost Columns",
        "Quantity and cost fields are converted into usable numbers where possible.",
        "Commas, INR, and rupee symbols are removed before conversion. Blank source quantities remain blank in source columns and may become zero only in normalized columns.",
    ),
    ("Column Definition", "PO", "Amazon purchase order number.", "Copied from the source PO column as text."),
    ("Column Definition", "Vendor Code", "Vendor identifier from Amazon.", "Copied from Vendor or Vendor Code as text."),
    ("Column Definition", "ASIN", "Amazon Standard Identification Number.", "Copied from the source ASIN column as text."),
    ("Column Definition", "External ID", "External product identifier such as EAN, UPC, or ISBN.", "Copied as text to preserve leading zeroes where the source file provides them."),
    ("Column Definition", "External ID Type", "Type of external product identifier.", "Copied from the source file when available."),
    ("Column Definition", "Ship to location", "Amazon destination or ship-to location.", "Copied from the source file when available."),
    ("Column Definition", "Model Number", "Vendor or product model number.", "Copied from the source file as text."),
    ("Column Definition", "Title", "Product title or description.", "Copied from the source file."),
    ("Column Definition", "Backordered", "Backorder indicator from Amazon, if supplied.", "Copied from the source file when available."),
    ("Column Definition", "Availability", "Availability status from the PO item report.", "Copied from the source file."),
    ("Column Definition", "Window Type", "Type of PO delivery or availability window.", "Copied from the source file."),
    ("Column Definition", "Window Start", "Start date of the PO window.", "Parsed from the source as day/month/year and displayed as a date."),
    ("Column Definition", "Window End", "End date of the PO window.", "Parsed from the source as day/month/year and displayed as a date."),
    ("Column Definition", "Expected Date", "Expected date from Amazon.", "Parsed from the source as month/day/year and displayed as a date."),
    ("Column Definition", "Quantity Requested", "Quantity originally requested on the PO.", "Copied from the source and converted to a number when valid."),
    ("Column Definition", "Quantity Accepted", "Quantity accepted by the vendor or Amazon workflow.", "Copied from the source and converted to a number when valid."),
    ("Column Definition", "Quantity Received", "Quantity already received by Amazon.", "Copied from the source and converted to a number when valid."),
    ("Column Definition", "Quantity Outstanding", "Outstanding quantity shown in the Amazon source file.", "Copied from the source and converted to a number when valid."),
    ("Column Definition", "Quantity Received Normalized", "Received quantity used for open PO calculations.", "Uses 0 when Quantity Received is blank; otherwise uses Quantity Received."),
    ("Column Definition", "Quantity Outstanding Normalized", "Outstanding quantity in calculation-ready form.", "Uses 0 when Quantity Outstanding is blank; otherwise uses Quantity Outstanding."),
    ("Column Definition", "Open PO Qty - Source", "Open PO quantity according to the source outstanding field.", "Equals Quantity Outstanding Normalized."),
    ("Column Definition", "Open PO Qty - Derived", "Open PO quantity calculated from accepted and received quantities.", "Calculated as Quantity Accepted minus Quantity Received Normalized, with a minimum of 0."),
    ("Column Definition", "Open PO Qty - Final", "Final open PO quantity used for reporting.", "Uses source outstanding when it is present; otherwise uses the derived open PO quantity."),
    ("Column Definition", "Unit Cost", "Unit cost from the PO item report.", "Copied from the source and converted to a number when valid."),
    ("Column Definition", "Unit Cost Currency", "Currency used for Unit Cost.", "Uses detected currency from the source or INR when a cost is present and no currency is stated."),
    ("Column Definition", "Total Cost", "Total line cost from the PO item report.", "Copied from the source and converted to a number when valid."),
    ("Column Definition", "Total Cost Currency", "Currency used for Total Cost.", "Uses detected currency from the source or INR when a cost is present and no currency is stated."),
    ("Column Definition", "Open PO Value - Final", "Estimated value of remaining open PO quantity.", "Calculated as Open PO Qty - Final multiplied by Unit Cost when both values are available."),
    ("Column Definition", "Remarks", "Blank column reserved for team comments.", "Always left blank by the pipeline."),
)


def _team_row(row: NormalizedPoItemRow) -> list[Any]:
    return ["" if header == "Remarks" else row.data.get(header) for header in TEAM_WORKBOOK_HEADERS]


def _write_sheet(ws: Any, headers: Iterable[str], rows: Iterable[Iterable[Any]], table_name: str) -> None:
    headers = list(headers)
    _assert_unique_headers(headers, ws.title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    _style_sheet(ws, table_name)


def _style_sheet(ws: Any, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="1F1F1F")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    last_row = max(ws.max_row, 1)
    last_col = max(ws.max_column, 1)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = table_ref
    _autosize_columns(ws)


def _format_master_sheet(ws: Any) -> None:
    header_positions = {cell.value: cell.column for cell in ws[1]}
    for header in ("External ID", "External ID Type", "PO", "Vendor Code", "ASIN", "Model Number", "Ship to location"):
        _format_column(ws, header_positions, header, "@")
    for header in ("Window Start", "Window End", "Expected Date"):
        _format_column(ws, header_positions, header, "dd-mmm-yyyy")
    for header in (
        "Quantity Requested",
        "Quantity Accepted",
        "Quantity Received",
        "Quantity Outstanding",
        "Quantity Received Normalized",
        "Quantity Outstanding Normalized",
        "Open PO Qty - Source",
        "Open PO Qty - Derived",
        "Open PO Qty - Final",
    ):
        _format_column(ws, header_positions, header, "0")
    for header in ("Unit Cost", "Total Cost", "Open PO Value - Final"):
        _format_column(ws, header_positions, header, "#,##0.00")
    if "Title" in header_positions:
        col = header_positions["Title"]
        ws.column_dimensions[get_column_letter(col)].width = 48
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].alignment = Alignment(wrap_text=True, vertical="top")
    if "Remarks" in header_positions:
        ws.column_dimensions[get_column_letter(header_positions["Remarks"])].width = 28


def _format_guide_sheet(ws: Any) -> None:
    widths = {
        "A": 24,
        "B": 34,
        "C": 62,
        "D": 70,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _format_column(ws: Any, header_positions: dict[str, int], header: str, number_format: str) -> None:
    col = header_positions.get(header)
    if col is None:
        return
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        row[0].number_format = number_format


def _autosize_columns(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 55)


def _assert_unique_headers(headers: list[str], sheet_name: str) -> None:
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise OutputWriteError(f"Duplicate output headers in {sheet_name}: {duplicates}")
