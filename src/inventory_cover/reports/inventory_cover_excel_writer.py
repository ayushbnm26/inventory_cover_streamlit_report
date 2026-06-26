"""Excel writers for the Inventory Cover team report and backend audit workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from inventory_cover.calculations import inventory_cover_formulas as F
from inventory_cover.exceptions import OutputWriteError
from inventory_cover.inventory_cover_schemas import (
    BUCKET_SHEET_NAMES,
    CALCULATION_AUDIT_HEADERS,
    COVER_BUCKETS,
    DATE_COLUMNS,
    DECIMAL_COLUMNS,
    FORMULA_GUIDE_HEADERS,
    INTEGER_COLUMNS,
    MASTER_BACKEND_HEADERS,
    RUN_METADATA_HEADERS,
    SOURCE_ROW_TRACE_HEADERS,
    SOURCE_SUMMARY_HEADERS,
    TEAM_REPORT_HEADERS,
    VALIDATION_ISSUE_HEADERS,
    InventoryCoverValidationIssue,
    ProductCoverRow,
    SourceSummaryRecord,
    SourceTraceRecord,
)
from inventory_cover.calculations.inventory_cover_builder import calculation_audit_row


HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

BUCKET_FILLS: dict[str, PatternFill] = {
    "Critical": PatternFill("solid", fgColor="F8CBAD"),
    "High Risk": PatternFill("solid", fgColor="FCE4D6"),
    "Watch": PatternFill("solid", fgColor="FFF2CC"),
    "Near Target": PatternFill("solid", fgColor="E2EFDA"),
    "Healthy": PatternFill("solid", fgColor="D9E1F2"),
    "No Sales": PatternFill("solid", fgColor="EDEDED"),
}


def write_team_workbook(
    output_path: Path,
    run_id: str,
    products: list[ProductCoverRow],
    summaries: list[SourceSummaryRecord],
    guide_rows: list[list[str]],
    config: Any,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        report_ws = wb.active
        report_ws.title = "Inventory_Cover_Report"
        _write_report_sheet(report_ws, products, "InventoryCoverReportTable", config, with_formulas=True)

        for bucket in COVER_BUCKETS:
            sheet_name = BUCKET_SHEET_NAMES[bucket]
            ws = wb.create_sheet(sheet_name)
            bucket_products = [p for p in products if p.cover_bucket == bucket]
            _write_annexure_sheet(ws, bucket_products, f"{_safe_table(sheet_name)}Table")

        guide_ws = wb.create_sheet("Formula_Guide")
        _write_plain_sheet(guide_ws, FORMULA_GUIDE_HEADERS, guide_rows, "InventoryCoverFormulaGuideTable")
        _format_guide_sheet(guide_ws)

        summary_ws = wb.create_sheet("Source_Summary")
        _write_plain_sheet(
            summary_ws,
            SOURCE_SUMMARY_HEADERS,
            [record.as_row() for record in summaries],
            "InventoryCoverSourceSummaryTable",
        )
        _format_long_text_columns(summary_ws, ("Warnings", "Source Latest Path", "Copied Run Path"))

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:  # noqa: BLE001
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write team workbook {output_path}: {exc}") from exc


def write_backend_workbook(
    output_path: Path,
    run_id: str,
    products: list[ProductCoverRow],
    traces: list[SourceTraceRecord],
    summaries: list[SourceSummaryRecord],
    issues: list[InventoryCoverValidationIssue],
    metadata_kv: list[tuple[str, Any]],
    guide_rows: list[list[str]],
    config: Any,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_suffix(output_path.suffix + ".tmp.xlsx")
    try:
        wb = Workbook()
        master_ws = wb.active
        master_ws.title = "Inventory_Cover_Master"
        _write_master_backend_sheet(master_ws, products, run_id, config)

        trace_ws = wb.create_sheet("Source_Row_Trace")
        _write_plain_sheet(
            trace_ws,
            SOURCE_ROW_TRACE_HEADERS,
            [record.as_row() for record in traces],
            "InventoryCoverSourceRowTraceTable",
        )

        summary_ws = wb.create_sheet("Source_Summary")
        _write_plain_sheet(
            summary_ws,
            SOURCE_SUMMARY_HEADERS,
            [record.as_row() for record in summaries],
            "InventoryCoverBackendSourceSummaryTable",
        )
        _format_long_text_columns(summary_ws, ("Warnings",))

        issues_ws = wb.create_sheet("Validation_Issues")
        _write_plain_sheet(
            issues_ws,
            VALIDATION_ISSUE_HEADERS,
            [issue.as_row() for issue in issues],
            "InventoryCoverValidationIssuesTable",
        )
        _format_long_text_columns(issues_ws, ("Issue Detail", "Action Taken", "Raw Value"))

        audit_ws = wb.create_sheet("Calculation_Audit")
        _write_plain_sheet(
            audit_ws,
            CALCULATION_AUDIT_HEADERS,
            [calculation_audit_row(product, run_id, config) for product in products],
            "InventoryCoverCalculationAuditTable",
        )
        _format_long_text_columns(audit_ws, ("DRR Formula", "Cover Formula Used", "Bucket Formula Used", "Warnings"))

        guide_ws = wb.create_sheet("Formula_Guide")
        _write_plain_sheet(guide_ws, FORMULA_GUIDE_HEADERS, guide_rows, "InventoryCoverBackendFormulaGuideTable")
        _format_guide_sheet(guide_ws)

        metadata_ws = wb.create_sheet("Run_Metadata")
        _write_plain_sheet(
            metadata_ws,
            RUN_METADATA_HEADERS,
            [[key, value] for key, value in metadata_kv],
            "InventoryCoverRunMetadataTable",
        )
        _format_long_text_columns(metadata_ws, ("Value",))

        wb.save(temp_path)
        temp_path.replace(output_path)
    except Exception as exc:  # noqa: BLE001
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)
        raise OutputWriteError(f"Could not write backend workbook {output_path}: {exc}") from exc


def _write_report_sheet(
    ws: Any,
    products: list[ProductCoverRow],
    table_name: str,
    config: Any,
    with_formulas: bool,
) -> None:
    headers = list(TEAM_REPORT_HEADERS)
    _assert_unique_headers(headers, ws.title)
    ws.append(headers)
    window = int(config.sales_window_days)
    default_target = float(config.default_target_doh)
    for product in products:
        row_cells: list[Any] = []
        for header in headers:
            formula = F.formula_for(header, window, default_target) if with_formulas else None
            if formula is not None:
                row_cells.append(formula)
            else:
                row_cells.append(product.team_value(header))
        ws.append(row_cells)
    _style_table(ws, table_name)
    _format_report_columns(ws)
    _apply_bucket_conditional(ws, products)


def _write_annexure_sheet(ws: Any, products: list[ProductCoverRow], table_name: str) -> None:
    headers = list(TEAM_REPORT_HEADERS)
    ws.append(headers)
    if not products:
        note_row = ["No records in this bucket for this run."] + [None] * (len(headers) - 1)
        ws.append(note_row)
    else:
        for product in products:
            ws.append([product.team_value(header) for header in headers])
    _style_table(ws, table_name)
    _format_report_columns(ws)


def _write_master_backend_sheet(ws: Any, products: list[ProductCoverRow], run_id: str, config: Any) -> None:
    headers = list(MASTER_BACKEND_HEADERS)
    _assert_unique_headers(headers, ws.title)
    ws.append(headers)
    for product in products:
        base = [product.team_value(header) for header in TEAM_REPORT_HEADERS]
        ws.append(base + product.backend_extra_values(run_id))
    _style_table(ws, "InventoryCoverMasterTable")
    _format_report_columns(ws)


def _write_plain_sheet(ws: Any, headers: Iterable[str], rows: Iterable[Iterable[Any]], table_name: str) -> None:
    headers = list(headers)
    _assert_unique_headers(headers, ws.title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    _style_table(ws, table_name)


def _style_table(ws: Any, table_name: str) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    last_row = max(ws.max_row, 1)
    last_col = max(ws.max_column, 1)
    table_ref = f"A1:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = table_ref
    _autosize_columns(ws)


def _format_report_columns(ws: Any) -> None:
    positions = {cell.value: cell.column for cell in ws[1]}
    for header in DATE_COLUMNS:
        _format_column(ws, positions, header, "dd-mmm-yyyy")
    for header in INTEGER_COLUMNS:
        _format_column(ws, positions, header, "#,##0")
    for header in DECIMAL_COLUMNS:
        _format_column(ws, positions, header, "0.00")
    _format_long_text_columns(ws, ("Cover Alert", "Data Quality Flag"))


def _apply_bucket_conditional(ws: Any, products: list[ProductCoverRow]) -> None:
    positions = {cell.value: cell.column for cell in ws[1]}
    bucket_col = positions.get("Cover Bucket")
    if bucket_col is None:
        return
    for index, product in enumerate(products, start=2):
        fill = BUCKET_FILLS.get(product.cover_bucket)
        if fill is None:
            continue
        ws.cell(row=index, column=bucket_col).fill = fill


def _format_column(ws: Any, positions: dict[Any, int], header: str, number_format: str) -> None:
    col = positions.get(header)
    if col is None:
        return
    for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
        row[0].number_format = number_format


def _format_long_text_columns(ws: Any, headers: tuple[str, ...]) -> None:
    positions = {cell.value: cell.column for cell in ws[1]}
    for header in headers:
        col = positions.get(header)
        if col is None:
            continue
        ws.column_dimensions[get_column_letter(col)].width = 42
        for cell in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            cell[0].alignment = Alignment(wrap_text=True, vertical="top")


def _format_guide_sheet(ws: Any) -> None:
    _format_long_text_columns(ws, ("Source / Formula", "Explanation", "Operational Meaning"))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 50


def _autosize_columns(ws: Any) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 10), 48)


def _assert_unique_headers(headers: list[str], sheet_name: str) -> None:
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise OutputWriteError(f"Duplicate output headers in {sheet_name}: {duplicates}")


def _safe_table(name: str) -> str:
    return "".join(ch for ch in name if ch.isalnum())


def build_formula_guide_rows(
    run_id: str,
    run_timestamp: str,
    config: Any,
    summaries: list[SourceSummaryRecord],
    product_count: int,
) -> list[list[str]]:
    """Build the Formula_Guide rows so the team can read the report unaided."""

    window = int(config.sales_window_days)
    default_target = float(config.default_target_doh)
    by_type = {summary.source_type: summary for summary in summaries}

    def src(source_type: str) -> str:
        summary = by_type.get(source_type)
        return summary.source_latest_path if summary else "Not configured"

    def period(source_type: str) -> str:
        summary = by_type.get(source_type)
        if not summary:
            return "Not available"
        start = summary.report_period_start or "n/a"
        end = summary.report_period_end or "n/a"
        return f"{start} -> {end}"

    rows: list[list[str]] = [
        ["Overview", "Purpose of report",
         "Final product-level inventory cover.",
         "Consolidates latest Sales, Inventory, B2B dispatch, and open PO data into one cover view.",
         "Tells the team how many days of supply each product has and where to act."],
        ["Overview", "Run date and run ID",
         f"Run ID {run_id}; generated {run_timestamp}.",
         "Each run is timestamped and traceable; inputs are copied into the run folder.",
         "Use the run ID when discussing a specific report version."],
        ["Sources", "Source workbooks used",
         f"Sales: {src('Sales')}; Inventory: {src('Inventory')}; B2B: {src('B2B Dispatch')}; "
         f"PO: {src('PO Items')}; ASIN Master: {src('ASIN Master')}.",
         "The engine reads the latest backend artifacts of the source pipelines.",
         "Source pipelines stay independent; this engine only consumes their outputs."],
        ["Sources", "Source sheet names",
         "Sales_Master, Inventory_Master, B2B_Dispatch_Master, PO_Items_Master, ASIN_Master.",
         "Stable sheet names form the interface contract.",
         "If a source sheet is renamed, update the engine's contract, not the calculations."],
        ["Sources", "Report dates from Sales",
         f"Sales period: {period('Sales')}.",
         "Parsed from Viewing Range Start / Viewing Range End in Sales_Master.",
         "Defines the sales window used for the Daily Run Rate."],
        ["Sources", "Report dates from Inventory",
         f"Inventory period: {period('Inventory')}.",
         "Parsed from Viewing Range End / Report Updated Date in Inventory_Master.",
         "Used for inventory freshness checks."],
        ["Sources", "Dispatch date window",
         "Own in-transit uses B2B dispatch rows flagged Included In Lookback Window.",
         "Pipeline 2 already filters the recent dispatch window.",
         "Captures stock you have shipped but Amazon has not yet received."],
        ["Sources", "PO date / window availability",
         f"PO period: {period('PO Items')}.",
         "Open PO often has no true PO date; only Window/Expected dates may exist and are not used to filter.",
         "Open PO is summed by product, not date-filtered, to avoid dropping valid PO quantity."],
        ["Identity", "Product identity logic",
         "Primary key ASIN; fallback Model Number / SKU. Product Title is only used for trace.",
         "Union of products across Inventory, Sales, PO, B2B, and ASIN Master.",
         "Every product that appears in any source is represented once."],
        ["Identity", "ASIN vs Model Number / SKU",
         "If ASIN is missing but Model Number exists, the row is kept and flagged Missing ASIN.",
         "A deterministic model->ASIN map merges model-only rows into their ASIN when unambiguous.",
         "Avoids duplicate product rows."],
        ["Policy", "Blank numeric values treated as zero",
         f"Policy: {config.blank_numeric_policy}.",
         "Blank Sales Units, On Hand, transit, and Open PO are treated as zero for calculation only; "
         "raw source values are preserved in backend audit sheets.",
         "Calculations never break on blank inputs."],
        ["Formula", "Sales Days",
         F.sales_days_formula(window),
         f"Sales Days = MIN({window}, period end - start + 1). If period is missing but sales exist, {window} is used.",
         "Caps the run-rate window so monthly data does not overstate daily demand."],
        ["Formula", "Daily Run Rate",
         F.daily_run_rate_formula(),
         "Daily Run Rate = Sales Units / Sales Days, divide-by-zero safe.",
         "Average units sold per day in the latest window."],
        ["Formula", "Current Stock DOH",
         F.current_stock_doh_formula(),
         "Sellable On Hand Units / Daily Run Rate; 'No Sales' when DRR is zero.",
         "Days of cover from on-hand stock alone."],
        ["Formula", "Stock + Amazon Transit DOH",
         F.stock_amazon_doh_formula(),
         "(On Hand + Amazon In-Transit) / Daily Run Rate.",
         "Cover including stock Amazon is already moving."],
        ["Formula", "Stock + Own Transit DOH",
         F.stock_own_doh_formula(),
         "(On Hand + Own In-Transit) / Daily Run Rate.",
         "Cover including your own dispatched stock."],
        ["Formula", "Total Transit DOH",
         F.total_transit_doh_formula(),
         "(On Hand + Amazon In-Transit + Own In-Transit) / Daily Run Rate.",
         "Cover including all in-transit stock."],
        ["Formula", "DOC Including Open PO",
         F.doc_including_open_po_formula(),
         "(On Hand + Open PO + Own In-Transit) / Daily Run Rate.",
         "Cover counting confirmed open purchase orders."],
        ["Formula", "Total Supply Cover DOH",
         F.total_supply_cover_doh_formula(),
         "(On Hand + Open PO + Amazon In-Transit + Own In-Transit) / Daily Run Rate.",
         "Full supply cover; drives the Cover Bucket."],
        ["Formula", "Target DOH",
         F.target_doh_formula(default_target),
         f"Aligned DOH Target if available, else default {default_target:g}.",
         "Desired days of cover per product."],
        ["Formula", "Gap to Target Units",
         F.gap_to_target_units_formula(),
         "MAX(Target DOH * DRR - total supply units, 0).",
         "Units still needed to reach the target cover."],
        ["Thresholds", "Cover Bucket thresholds",
         "<5 Critical; 5-15 High Risk; 15-25 Watch; 25-30 Near Target; >30 Healthy.",
         "Boundaries: <5, >=5 & <15, >=15 & <25, >=25 & <=30, >30. Based on Total Supply Cover DOH.",
         "Single clear bucket per product for prioritisation."],
        ["Thresholds", "Cover Alert meanings",
         "Critical=Immediate action; High Risk=Urgent replenishment; Watch=Plan replenishment; "
         "Near Target=Monitor; Healthy=No immediate action.",
         "Derived from the Cover Bucket.",
         "Plain-language next step for each product."],
        ["Handling", "No Sales handling",
         "When DRR is zero, DOH columns show 'No Sales' and bucket is 'No Sales'.",
         "Prevents divide-by-zero and misleading infinite cover.",
         "Surfaces products with stock but no recent sales."],
        ["Handling", "Data Quality Flag",
         "Computed in Python (joined with '; ').",
         "Values: OK, No Sales, Missing ASIN, Missing Inventory, Missing Sales, Missing Target DOH, "
         "Identifier Conflict, Calculation Warning.",
         "Quick read on row trustworthiness."],
        ["Handling", "Remarks column purpose",
         "Left blank intentionally.",
         "Free space for the team to add notes.",
         "Not used by any calculation."],
        ["Scope", "What this report does not do",
         "No forecasting, no automatic PO creation, no pricing, no lead-time modelling.",
         "It is a cover snapshot from the latest available source data.",
         "Use it to prioritise, not to auto-order."],
        ["Scope", "How to use the report operationally",
         f"Sort by Cover Bucket; act on Critical and High Risk first. {product_count} products in this run.",
         "Filter the Excel table or open the bucket annexure sheets.",
         "Drive replenishment and escalation decisions."],
    ]
    return rows
