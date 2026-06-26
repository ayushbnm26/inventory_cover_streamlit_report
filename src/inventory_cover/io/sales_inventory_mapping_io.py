"""Read optional ASIN/SKU mapping workbooks for Sales & Inventory enrichment."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from inventory_cover.exceptions import FileValidationError
from inventory_cover.sales_inventory_schemas import SalesInventoryMappingAuditRecord
from inventory_cover.utils.text_cleaning import clean_text, is_blank, normalize_header


MAPPING_FIELDS: tuple[str, ...] = ("ASIN", "SKU", "Master SKU")
MAPPING_HEADER_LOOKUP: dict[str, str] = {
    normalize_header("ASIN"): "ASIN",
    normalize_header("SKU"): "SKU",
    normalize_header("Model Number"): "SKU",
    normalize_header("Master SKU"): "Master SKU",
    normalize_header("MasterSKU"): "Master SKU",
}


@dataclass(frozen=True)
class MappingMatch:
    value: str | None
    status: str

    @property
    def found(self) -> bool:
        return self.status == "FOUND" and bool(self.value)


@dataclass(frozen=True)
class SalesInventoryMappingLookup:
    asin_to_model: dict[str, str]
    model_to_asin: dict[str, str]
    ambiguous_asin_keys: set[str]
    ambiguous_model_keys: set[str]

    def model_for_asin(self, asin: str) -> MappingMatch:
        key = _key(asin)
        if not key:
            return MappingMatch(None, "MISSING_KEY")
        if key in self.ambiguous_asin_keys:
            return MappingMatch(None, "AMBIGUOUS")
        value = self.asin_to_model.get(key)
        return MappingMatch(value, "FOUND" if value else "NOT_FOUND")

    def asin_for_model(self, model_number: str) -> MappingMatch:
        key = _key(model_number)
        if not key:
            return MappingMatch(None, "MISSING_KEY")
        if key in self.ambiguous_model_keys:
            return MappingMatch(None, "AMBIGUOUS")
        value = self.model_to_asin.get(key)
        return MappingMatch(value, "FOUND" if value else "NOT_FOUND")


@dataclass(frozen=True)
class SalesInventoryMappingReadResult:
    lookup: SalesInventoryMappingLookup
    audit: SalesInventoryMappingAuditRecord


def read_sales_inventory_mapping_workbook(
    path: Path,
    run_id: str,
    copied_run_path: Path | None = None,
) -> SalesInventoryMappingReadResult:
    """Read a mapping workbook into deterministic lookup dictionaries."""

    audit = SalesInventoryMappingAuditRecord(
        run_id=run_id,
        mapping_file=path.name,
        full_path=str(path),
        copied_run_path=str(copied_run_path or ""),
    )
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        audit.status = "FAILED"
        audit.notes = f"Could not open mapping workbook: {exc}"
        raise FileValidationError("Could not open mapping workbook", details=str(exc)) from exc

    try:
        for ws in wb.worksheets:
            reset = getattr(ws, "reset_dimensions", None)
            if callable(reset):
                reset()
        detection = _detect_mapping_header(wb.worksheets)
        if detection is None:
            audit.status = "FAILED"
            audit.notes = "No recognizable ASIN/SKU mapping header row found."
            raise FileValidationError("Mapping header row not found", details=audit.notes)

        sheet_name, header_row, mapping = detection
        ws = wb[sheet_name]
        reset = getattr(ws, "reset_dimensions", None)
        if callable(reset):
            reset()

        audit.source_sheet = sheet_name
        audit.header_row_found = header_row

        asin_to_models: dict[str, set[str]] = {}
        model_to_asins: dict[str, set[str]] = {}
        rows_loaded = 0
        rows_scanned = 0

        for source_row, cells in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            cell_list = list(cells)
            if all(is_blank(cell.value) for cell in cell_list):
                continue
            rows_scanned += 1
            asin = _cell_text(cell_list, mapping.get("ASIN"))
            sku = _cell_text(cell_list, mapping.get("SKU"))
            master_sku = _cell_text(cell_list, mapping.get("Master SKU"))
            model = sku or master_sku
            if not asin or not model:
                continue
            rows_loaded += 1
            asin_to_models.setdefault(_key(asin), set()).add(model)
            model_to_asins.setdefault(_key(model), set()).add(asin)
            if master_sku:
                model_to_asins.setdefault(_key(master_sku), set()).add(asin)

        asin_to_model = {key: next(iter(values)) for key, values in asin_to_models.items() if len(values) == 1}
        model_to_asin = {key: next(iter(values)) for key, values in model_to_asins.items() if len(values) == 1}
        ambiguous_asin_keys = {key for key, values in asin_to_models.items() if len(values) > 1}
        ambiguous_model_keys = {key for key, values in model_to_asins.items() if len(values) > 1}

        audit.rows_scanned = rows_scanned
        audit.rows_loaded = rows_loaded
        audit.unique_asin_keys = len(asin_to_model)
        audit.unique_sku_keys = len(model_to_asin)
        audit.ambiguous_asin_keys = len(ambiguous_asin_keys)
        audit.ambiguous_sku_keys = len(ambiguous_model_keys)
        audit.status = "SUCCESS_WITH_WARNINGS" if ambiguous_asin_keys or ambiguous_model_keys else "SUCCESS"
        audit.notes = (
            "Ambiguous keys were skipped; no mapping guesses were made."
            if ambiguous_asin_keys or ambiguous_model_keys
            else "Mapping lookup loaded successfully."
        )
        return SalesInventoryMappingReadResult(
            lookup=SalesInventoryMappingLookup(
                asin_to_model=asin_to_model,
                model_to_asin=model_to_asin,
                ambiguous_asin_keys=ambiguous_asin_keys,
                ambiguous_model_keys=ambiguous_model_keys,
            ),
            audit=audit,
        )
    finally:
        wb.close()


def _detect_mapping_header(worksheets: list[Any], max_scan_rows: int = 20) -> tuple[str, int, dict[str, int]] | None:
    best: tuple[str, int, dict[str, int], int] | None = None
    for ws in worksheets:
        for row_number, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
            start=1,
        ):
            mapping: dict[str, int] = {}
            for idx, value in enumerate(row):
                canonical = MAPPING_HEADER_LOOKUP.get(normalize_header(value))
                if canonical and canonical not in mapping:
                    mapping[canonical] = idx
            score = len(set(mapping).intersection(MAPPING_FIELDS))
            if score < 2 or "ASIN" not in mapping:
                continue
            if best is None or score > best[3]:
                best = (ws.title, row_number, mapping, score)
    if best is None:
        return None
    return best[0], best[1], best[2]


def _cell_text(cells: list[Any], idx: int | None) -> str:
    if idx is None or idx >= len(cells):
        return ""
    cell = cells[idx]
    return clean_text(cell.value, getattr(cell, "number_format", "") or "")


def _key(value: str) -> str:
    return str(value or "").strip().upper()
