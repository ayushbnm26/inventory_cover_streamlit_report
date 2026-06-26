"""Read latest backend artifacts produced by the source pipelines.

This module only knows the stable backend workbook interface contract (sheet
names and column headers). It does not import or call any source-pipeline
reading logic, keeping the calculation engine loosely coupled.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from inventory_cover.utils.text_cleaning import normalize_header


@dataclass
class LoadedSheet:
    """A single backend master sheet loaded into header-keyed dict rows."""

    source_type: str
    path: Path
    exists: bool = False
    sheet_name: str = ""
    headers: list[str] = field(default_factory=list)
    rows: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def get(self, row: dict[str, Any], header: str) -> Any:
        return row.get(header)


def load_backend_sheet(
    source_type: str,
    path: Path,
    sheet_name: str,
) -> LoadedSheet:
    """Load one backend workbook sheet by name, tolerating absence gracefully."""

    loaded = LoadedSheet(source_type=source_type, path=path)
    if not path.exists():
        loaded.warnings.append(f"Backend workbook not found: {path}")
        return loaded

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - report and continue
        loaded.warnings.append(f"Could not open workbook {path}: {exc}")
        return loaded

    try:
        loaded.exists = True
        target = _resolve_sheet(wb, sheet_name)
        if target is None:
            loaded.warnings.append(
                f"Expected sheet '{sheet_name}' not found in {path.name}; "
                f"available sheets: {', '.join(wb.sheetnames)}"
            )
            return loaded

        loaded.sheet_name = target.title
        rows_iter = target.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            loaded.warnings.append(f"Sheet '{target.title}' in {path.name} is empty.")
            return loaded

        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        loaded.headers = headers
        for raw in rows_iter:
            if raw is None:
                continue
            if all(cell is None or str(cell).strip() == "" for cell in raw):
                continue
            row: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = raw[index] if index < len(raw) else None
            loaded.rows.append(row)
        return loaded
    finally:
        wb.close()


def _resolve_sheet(wb: Any, sheet_name: str) -> Any:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    target_norm = normalize_header(sheet_name)
    for name in wb.sheetnames:
        if normalize_header(name) == target_norm:
            return wb[name]
    return None


def first_present_column(headers: list[str], candidates: tuple[str, ...]) -> str | None:
    """Return the first candidate header present in the sheet (tolerant match)."""

    normalized = {normalize_header(header): header for header in headers}
    for candidate in candidates:
        key = normalize_header(candidate)
        if key in normalized:
            return normalized[key]
    return None
