"""Excel reading, header detection, and metadata parsing for Vendor Central exports."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.read_only import ReadOnlyCell

from inventory_cover.config import SalesInventoryPipelineConfig
from inventory_cover.exceptions import FileValidationError
from inventory_cover.sales_inventory_schemas import (
    INVENTORY_MINIMUM_HEADERS,
    INVENTORY_REPORT_TYPE,
    INVENTORY_SOURCE_COLUMNS,
    SALES_MINIMUM_HEADERS,
    SALES_REPORT_TYPE,
    SALES_SOURCE_COLUMNS,
    ReportType,
    RawSalesInventoryRow,
    SalesInventoryCellValue,
    SalesInventoryWorkbookReadResult,
)
from inventory_cover.utils.date_parsing import parse_vendor_central_date
from inventory_cover.utils.text_cleaning import is_blank, normalize_header


@dataclass(frozen=True)
class SalesInventoryHeaderDetection:
    sheet_name: str
    header_row: int
    mapping: dict[str, int]
    source_headers: tuple[str, ...]
    missing_minimum_headers: tuple[str, ...]
    score: int


def build_sales_inventory_header_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for field in set(SALES_SOURCE_COLUMNS + INVENTORY_SOURCE_COLUMNS):
        variants = _header_variants(field)
        for value in (field, *variants):
            normalized = normalize_header(value)
            if normalized:
                lookup[normalized] = field
    return lookup


def map_sales_inventory_headers(header_values: list[Any]) -> dict[str, int]:
    """Map Vendor Central source headers to canonical field names."""

    mapping: dict[str, int] = {}
    for idx, value in enumerate(header_values):
        canonical = SALES_INVENTORY_HEADER_LOOKUP.get(normalize_header(value))
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


def detect_sales_inventory_header_row_in_sheet(
    ws: Any,
    report_type: ReportType,
    max_scan_rows: int = 20,
) -> SalesInventoryHeaderDetection | None:
    """Find the most likely Vendor Central header row in one worksheet."""

    expected = _expected_columns(report_type)
    minimum = _minimum_columns(report_type)
    best: SalesInventoryHeaderDetection | None = None
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        header_values = list(row)
        mapping = map_sales_inventory_headers(header_values)
        score = len(set(mapping).intersection(expected))
        if score == 0:
            continue
        missing = tuple(field for field in minimum if field not in mapping)
        source_headers = tuple(str(value).strip() for value in header_values if not is_blank(value))
        detection = SalesInventoryHeaderDetection(
            sheet_name=ws.title,
            header_row=row_number,
            mapping=mapping,
            source_headers=source_headers,
            missing_minimum_headers=missing,
            score=score,
        )
        if best is None or detection.score > best.score:
            best = detection
    return best


def read_sales_inventory_workbook(
    path: Path,
    config: SalesInventoryPipelineConfig,
    report_type: ReportType,
) -> SalesInventoryWorkbookReadResult:
    """Read one sales or inventory workbook into raw row records."""

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise FileValidationError("Could not open workbook", details=str(exc)) from exc

    try:
        for ws in wb.worksheets:
            _reset_dimensions_if_needed(ws)
        detections = [
            detection
            for detection in (
                detect_sales_inventory_header_row_in_sheet(
                    ws,
                    report_type=report_type,
                    max_scan_rows=config.header_scan_rows,
                )
                for ws in wb.worksheets
            )
            if detection is not None
        ]
        if not detections:
            raise FileValidationError(
                "Header row not found",
                details=f"No recognizable {report_type.lower()} Vendor Central headers.",
            )
        best = max(detections, key=lambda detection: detection.score)
        if best.missing_minimum_headers:
            raise FileValidationError(
                "Minimum identifying headers missing",
                missing_headers=list(best.missing_minimum_headers),
            )

        ws = wb[best.sheet_name]
        _reset_dimensions_if_needed(ws)
        expected = _expected_columns(report_type)
        raw_metadata_text = _raw_metadata_text(ws, best.header_row)
        metadata = parse_vendor_central_metadata(raw_metadata_text)
        rows, blank_skipped = _read_data_rows(path, ws, best, report_type, expected, metadata)

        found_expected = tuple(field for field in expected if field in best.mapping)
        missing_expected = tuple(field for field in expected if field not in best.mapping)
        extra_headers = tuple(
            header
            for header in best.source_headers
            if SALES_INVENTORY_HEADER_LOOKUP.get(normalize_header(header)) is None
        )

        return SalesInventoryWorkbookReadResult(
            source_path=path,
            report_type=report_type,
            sheet_name=best.sheet_name,
            header_row=best.header_row,
            rows=rows,
            metadata=metadata,
            raw_metadata_text=raw_metadata_text,
            expected_columns=expected,
            found_expected_columns=found_expected,
            missing_expected_columns=missing_expected,
            extra_source_columns=extra_headers,
            rows_blank_skipped=blank_skipped,
        )
    finally:
        wb.close()


def parse_vendor_central_metadata(text: str) -> dict[str, Any]:
    """Parse the metadata/filter line that Amazon places above the table header."""

    metadata: dict[str, Any] = {"Raw Metadata Text": text or ""}
    if not text:
        return metadata
    for key, value in re.findall(r"([^=\[\]]+)=\[(.*?)\]", text):
        clean_key = " ".join(str(key).strip().split())
        metadata[clean_key] = value.strip()

    viewing_range = str(metadata.get("Viewing Range") or "").strip()
    if viewing_range:
        parts = re.split(r"\s+-\s+", viewing_range, maxsplit=1)
        if len(parts) == 2:
            start = parse_vendor_central_date(parts[0])
            end = parse_vendor_central_date(parts[1])
            metadata["Viewing Range Start"] = start.value
            metadata["Viewing Range End"] = end.value

    report_updated = metadata.get("Report Updated")
    if report_updated:
        parsed = parse_vendor_central_date(report_updated)
        metadata["Report Updated Date"] = parsed.value

    if "Countries" in metadata and "Country" not in metadata:
        metadata["Country"] = metadata["Countries"]
    return metadata


def _read_data_rows(
    path: Path,
    ws: Any,
    detection: SalesInventoryHeaderDetection,
    report_type: ReportType,
    expected: tuple[str, ...],
    metadata: dict[str, Any],
) -> tuple[list[RawSalesInventoryRow], int]:
    rows: list[RawSalesInventoryRow] = []
    blank_skipped = 0
    for source_row, cells in enumerate(
        ws.iter_rows(min_row=detection.header_row + 1),
        start=detection.header_row + 1,
    ):
        cell_list = list(cells)
        if _is_blank_excel_row(cell_list):
            blank_skipped += 1
            continue
        values: dict[str, SalesInventoryCellValue] = {}
        for field in expected:
            idx = detection.mapping.get(field)
            if idx is None or idx >= len(cell_list):
                values[field] = SalesInventoryCellValue(value=None)
                continue
            cell = cell_list[idx]
            values[field] = SalesInventoryCellValue(
                value=cell.value,
                number_format=getattr(cell, "number_format", "") or "",
                data_type=getattr(cell, "data_type", "") or "",
            )
        rows.append(
            RawSalesInventoryRow(
                report_type=report_type,
                source_file=path.name,
                source_path=path,
                source_sheet=ws.title,
                source_row=source_row,
                values=values,
                metadata=metadata,
            )
        )
    return rows, blank_skipped


def _raw_metadata_text(ws: Any, header_row: int) -> str:
    values: list[str] = []
    if header_row <= 1:
        return ""
    for row in ws.iter_rows(min_row=1, max_row=header_row - 1, values_only=True):
        row_text = " ".join(str(value).strip() for value in row if not is_blank(value))
        if row_text:
            values.append(row_text)
    return " ".join(values)


def _expected_columns(report_type: ReportType) -> tuple[str, ...]:
    return SALES_SOURCE_COLUMNS if report_type == SALES_REPORT_TYPE else INVENTORY_SOURCE_COLUMNS


def _minimum_columns(report_type: ReportType) -> tuple[str, ...]:
    return SALES_MINIMUM_HEADERS if report_type == SALES_REPORT_TYPE else INVENTORY_MINIMUM_HEADERS


def _header_variants(field: str) -> tuple[str, ...]:
    variants: dict[str, tuple[str, ...]] = {
        "Child Vendor Code": ("Vendor Code", "Child Vendor", "ChildVendorCode"),
        "Product Title": ("Title", "Item Title", "Product Name"),
        "Parent ASIN": ("ParentASIN",),
        "Model Number": ("Model No", "Model", "Model #"),
        "Sellable On-Hand Inventory": ("Sellable On Hand Inventory",),
        "Unsellable On-Hand Inventory": ("Unsellable On Hand Inventory",),
        "Sellable On Hand Units": ("Sellable On-Hand Units",),
        "Unsellable On-Hand Units": ("Unsellable On Hand Units",),
        "Overall Vendor Lead Time (days)": ("Overall Vendor Lead Time Days",),
    }
    return variants.get(field, ())


SALES_INVENTORY_HEADER_LOOKUP = build_sales_inventory_header_lookup()


def _is_blank_excel_row(cells: list[ReadOnlyCell]) -> bool:
    return all(is_blank(cell.value) for cell in cells)


def _reset_dimensions_if_needed(ws: Any) -> None:
    reset_dimensions = getattr(ws, "reset_dimensions", None)
    if callable(reset_dimensions):
        reset_dimensions()
