"""Excel reading and header detection for B2B Dispatch Tracker workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.read_only import ReadOnlyCell

from inventory_cover.b2b_dispatch_schemas import (
    B2BCellValue,
    B2BWorkbookReadResult,
    B2BSheetAuditRecord,
    B2B_CRITICAL_FIELDS,
    B2B_HEADER_VARIANTS,
    B2B_SOURCE_FIELDS,
    B2B_TARGET_SHEETS,
    B2BTargetSheetSpec,
    RawB2BDispatchRow,
)
from inventory_cover.config import B2BDispatchPipelineConfig
from inventory_cover.exceptions import FileValidationError
from inventory_cover.utils.text_cleaning import is_blank, normalize_header


MIN_B2B_HEADER_SCORE = 4


@dataclass(frozen=True)
class B2BHeaderDetection:
    sheet_name: str
    header_row: int
    mapping: dict[str, int]
    missing_critical: list[str]
    score: int


def build_b2b_header_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, variants in B2B_HEADER_VARIANTS.items():
        for value in (canonical, *variants):
            normalized = normalize_header(value)
            if normalized:
                lookup[normalized] = canonical
    return lookup


B2B_HEADER_LOOKUP = build_b2b_header_lookup()


def map_b2b_headers(header_values: list[Any]) -> dict[str, int]:
    """Map dispatch tracker source headers to canonical field names."""

    mapping: dict[str, int] = {}
    for idx, value in enumerate(header_values):
        canonical = B2B_HEADER_LOOKUP.get(normalize_header(value))
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


def detect_b2b_header_row_in_sheet(ws: Any, max_scan_rows: int = 30) -> B2BHeaderDetection | None:
    """Find the most likely dispatch header row in one worksheet."""

    best: B2BHeaderDetection | None = None
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        mapping = map_b2b_headers(list(row))
        score = len(mapping)
        if score == 0:
            continue
        missing = [field for field in B2B_CRITICAL_FIELDS if field not in mapping]
        detection = B2BHeaderDetection(
            sheet_name=ws.title,
            header_row=row_number,
            mapping=mapping,
            missing_critical=missing,
            score=score,
        )
        if best is None or detection.score > best.score:
            best = detection
    if best is not None and best.score < MIN_B2B_HEADER_SCORE:
        return None
    return best


def read_b2b_dispatch_workbook(
    path: Path,
    config: B2BDispatchPipelineConfig,
    run_id: str,
) -> B2BWorkbookReadResult:
    """Read one dispatch tracker workbook into raw row records."""

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise FileValidationError("Could not open workbook", details=str(exc)) from exc

    try:
        rows: list[RawB2BDispatchRow] = []
        sheet_audit: list[B2BSheetAuditRecord] = []
        for spec in B2B_TARGET_SHEETS:
            audit = B2BSheetAuditRecord(
                run_id=run_id,
                source_file=path.name,
                expected_sheet_name=spec.expected_sheet_name,
                source_channel=spec.source_channel,
            )
            ws = _find_target_sheet(wb, spec)
            if ws is None:
                audit.status = "MISSING"
                audit.notes = "Expected target sheet was not found."
                sheet_audit.append(audit)
                continue

            audit.sheet_found = True
            audit.actual_sheet_name = ws.title
            detection = detect_b2b_header_row_in_sheet(ws, max_scan_rows=config.header_scan_rows)
            if detection is None:
                audit.status = "FAILED"
                audit.notes = "Header row could not be detected."
                sheet_audit.append(audit)
                continue
            audit.header_row_found = detection.header_row
            if detection.missing_critical:
                audit.status = "FAILED"
                audit.notes = "Critical headers missing: " + ", ".join(detection.missing_critical)
                sheet_audit.append(audit)
                continue

            sheet_rows = _read_sheet_rows(path, ws, detection, spec)
            audit.rows_scanned = len(sheet_rows)
            audit.status = "READ"
            rows.extend(sheet_rows)
            sheet_audit.append(audit)

        return B2BWorkbookReadResult(
            source_path=path,
            rows=rows,
            sheet_audit=sheet_audit,
        )
    finally:
        wb.close()


def _read_sheet_rows(
    path: Path,
    ws: Any,
    detection: B2BHeaderDetection,
    spec: B2BTargetSheetSpec,
) -> list[RawB2BDispatchRow]:
    rows: list[RawB2BDispatchRow] = []
    for source_row, cells in enumerate(
        ws.iter_rows(min_row=detection.header_row + 1),
        start=detection.header_row + 1,
    ):
        cell_list = list(cells)
        if _is_blank_excel_row(cell_list):
            continue
        values: dict[str, B2BCellValue] = {}
        for field in B2B_SOURCE_FIELDS:
            idx = detection.mapping.get(field)
            if idx is None or idx >= len(cell_list):
                values[field] = B2BCellValue(value=None)
                continue
            cell = cell_list[idx]
            values[field] = B2BCellValue(
                value=cell.value,
                number_format=getattr(cell, "number_format", "") or "",
                data_type=getattr(cell, "data_type", "") or "",
            )
        rows.append(
            RawB2BDispatchRow(
                source_file=path.name,
                source_path=path,
                source_sheet=ws.title,
                source_row=source_row,
                source_channel=spec.source_channel,
                values=values,
            )
        )
    return rows


def _find_target_sheet(wb: Any, spec: B2BTargetSheetSpec) -> Any | None:
    normalized_aliases = {normalize_header(alias) for alias in spec.aliases}
    for ws in wb.worksheets:
        if normalize_header(ws.title) in normalized_aliases:
            return ws
    return None


def _is_blank_excel_row(cells: list[ReadOnlyCell]) -> bool:
    return all(is_blank(cell.value) for cell in cells)
