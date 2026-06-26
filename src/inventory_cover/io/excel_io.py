"""Excel reading and header detection for Amazon PO Items exports."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.read_only import ReadOnlyCell

from inventory_cover.config import PipelineConfig
from inventory_cover.exceptions import FileValidationError
from inventory_cover.schemas import (
    CellValue,
    HEADER_VARIANTS,
    OPTIONAL_FIELDS,
    REQUIRED_FIELDS,
    RawPoItemRow,
    WorkbookReadResult,
)
from inventory_cover.utils.text_cleaning import is_blank, normalize_header


MIN_HEADER_SCORE = 8


@dataclass(frozen=True)
class HeaderDetection:
    sheet_name: str
    header_row: int
    mapping: dict[str, int]
    missing_required: list[str]
    score: int


def build_header_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, variants in HEADER_VARIANTS.items():
        for value in (canonical, *variants):
            normalized = normalize_header(value)
            if normalized:
                lookup[normalized] = canonical
    return lookup


HEADER_LOOKUP = build_header_lookup()


def map_headers(header_values: list[Any]) -> dict[str, int]:
    """Map source header positions to canonical field names."""

    mapping: dict[str, int] = {}
    for idx, value in enumerate(header_values):
        canonical = HEADER_LOOKUP.get(normalize_header(value))
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


def detect_header_row_in_sheet(ws: Any, max_scan_rows: int = 30) -> HeaderDetection | None:
    """Find the most likely header row in one worksheet."""

    best: HeaderDetection | None = None
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        header_values = list(row)
        mapping = map_headers(header_values)
        score = len(set(mapping).intersection(REQUIRED_FIELDS))
        if score == 0:
            continue
        missing = [field for field in REQUIRED_FIELDS if field not in mapping]
        detection = HeaderDetection(
            sheet_name=ws.title,
            header_row=row_number,
            mapping=mapping,
            missing_required=missing,
            score=score,
        )
        if best is None or detection.score > best.score:
            best = detection
    return best


def read_po_items_workbook(path: Path, config: PipelineConfig) -> WorkbookReadResult:
    """Read one workbook into raw row records after header validation."""

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several exception types for invalid files.
        raise FileValidationError("Could not open workbook", details=str(exc)) from exc

    try:
        detections = [
            detection
            for detection in (
                detect_header_row_in_sheet(ws, max_scan_rows=config.header_scan_rows)
                for ws in _candidate_sheets(wb)
            )
            if detection is not None
        ]
        if not detections:
            raise FileValidationError("Header row not found", details="No recognizable PO Items headers.")

        best = max(detections, key=lambda detection: detection.score)
        if best.score < MIN_HEADER_SCORE:
            raise FileValidationError(
                "Header row not found",
                details=f"Best header score was {best.score}; expected at least {MIN_HEADER_SCORE}.",
            )
        if best.missing_required:
            raise FileValidationError(
                "Critical headers missing",
                missing_headers=best.missing_required,
            )

        ws = wb[best.sheet_name]
        rows: list[RawPoItemRow] = []
        fields_to_read = tuple(REQUIRED_FIELDS) + tuple(OPTIONAL_FIELDS)
        for source_row, cells in enumerate(ws.iter_rows(min_row=best.header_row + 1), start=best.header_row + 1):
            cell_list = list(cells)
            if _is_blank_excel_row(cell_list):
                continue
            values: dict[str, CellValue] = {}
            for field in fields_to_read:
                idx = best.mapping.get(field)
                if idx is None or idx >= len(cell_list):
                    values[field] = CellValue(value=None)
                    continue
                cell = cell_list[idx]
                values[field] = CellValue(
                    value=cell.value,
                    number_format=getattr(cell, "number_format", "") or "",
                    data_type=getattr(cell, "data_type", "") or "",
                )
            rows.append(
                RawPoItemRow(
                    source_file=path.name,
                    source_path=path,
                    source_sheet=best.sheet_name,
                    source_row=source_row,
                    values=values,
                )
            )
        return WorkbookReadResult(
            source_path=path,
            sheet_name=best.sheet_name,
            header_row=best.header_row,
            rows=rows,
        )
    finally:
        wb.close()


def _candidate_sheets(wb: Any) -> list[Any]:
    preferred = [ws for ws in wb.worksheets if normalize_header(ws.title) == "purchaseorderitems"]
    others = [ws for ws in wb.worksheets if ws not in preferred]
    return preferred + others


def _is_blank_excel_row(cells: list[ReadOnlyCell]) -> bool:
    return all(is_blank(cell.value) for cell in cells)
