from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from inventory_cover.config import B2BDispatchPipelineConfig
from inventory_cover.exceptions import CatastrophicPipelineError
from inventory_cover.io.b2b_dispatch_excel_io import (
    detect_b2b_header_row_in_sheet,
    read_b2b_dispatch_workbook,
)
from inventory_cover.pipelines.b2b_dispatch_pipeline import B2BDispatchPipeline


AS_OF_DATE = date(2026, 6, 26)
RK_SHEET = "RK PO 007GK"
CLICK_SHEET = "CLICKTECK DISPATCH "
ETRADE_SHEET = "ETRADE DISPATCH "

RK_HEADERS = [
    "APPOINTMENT ID",
    "INVOICE NO",
    "BOXES",
    "PO",
    "LOC.",
    "ASIN",
    "PO+ASIN",
    "SKU",
    "PO DATE",
    "PO QTY",
    "Dispatch Qty",
    "UNIT VALUE",
    "TOTAL VALUE",
    "DATE",
    "LOCATION",
]

CLICK_HEADERS = [
    "Appointment ID",
    "INVOICE NO",
    "BOXES",
    "PO",
    "",
    "ASIN",
    "PO+ASIN",
    "MODEL NAME",
    "PO DATE",
    "PO QTY",
    "Dispatch Qty",
    "UNIT PRICE",
    "PO DISPATCH VALUE",
    "Date",
    "Location",
]

ETRADE_HEADERS = [
    "APPOINTMENT ID",
    "INVOICE NO",
    "BOXES",
    "PO",
    "Ship to location",
    "ASIN",
    "PO+ASIN",
    "MODEL NAME",
    "PO DATE",
    "PO QTY",
    "Dispatch Qty",
    "UNIT VALUE",
    "TOTAL VALUE",
    "Date",
    "Location",
]


def test_header_detection_works_when_headers_are_on_row_3() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = RK_SHEET
    ws.append(["summary"])
    ws.append(["generated"])
    ws.append(RK_HEADERS)

    detection = detect_b2b_header_row_in_sheet(ws)

    assert detection is not None
    assert detection.header_row == 3
    assert detection.missing_critical == []


def test_header_detection_works_when_headers_are_on_row_4() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = ETRADE_SHEET
    ws.append(["summary"])
    ws.append(["generated"])
    ws.append(["formula row"])
    ws.append(ETRADE_HEADERS)

    detection = detect_b2b_header_row_in_sheet(ws)

    assert detection is not None
    assert detection.header_row == 4
    assert detection.missing_critical == []


def test_only_rk_clicktech_and_etrade_sheets_are_processed(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={
            RK_SHEET: [_row(po="PO-RK", asin="ASIN-RK")],
            CLICK_SHEET: [_row(po="PO-CLICK", asin="ASIN-CLICK")],
            ETRADE_SHEET: [_row(po="PO-ETRADE", asin="ASIN-ETRADE")],
        },
        include_non_target=True,
    )

    result = B2BDispatchPipeline(_config(tmp_path, input_dir)).run()
    records = _master_records(result.backend_output_file)

    assert {record["Source Channel"] for record in records} == {"RK", "CLICKTECH", "ETRADE"}
    assert {record["Source Sheet"] for record in records} == {RK_SHEET, CLICK_SHEET, ETRADE_SHEET}


def test_non_target_sheets_are_ignored(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={RK_SHEET: [_row()]},
        include_sheets=(RK_SHEET,),
        include_non_target=True,
    )

    result = B2BDispatchPipeline(_config(tmp_path, input_dir, allow_missing=True)).run()
    records = _master_records(result.backend_output_file)

    assert len(records) == 1
    assert records[0]["Source Sheet"] == RK_SHEET


def test_date_filtering_keeps_today_and_yesterday_only(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={
            RK_SHEET: [
                _row(po="PO-TODAY", dispatch_date="26-06-2026"),
                _row(po="PO-YESTERDAY", dispatch_date="25/06/2026"),
                _row(po="PO-OLD", dispatch_date="24-06-2026"),
            ]
        },
        include_sheets=(RK_SHEET,),
    )

    result = B2BDispatchPipeline(_config(tmp_path, input_dir, allow_missing=True)).run()
    records = _master_records(result.backend_output_file)
    issue_types = _issue_types(result.backend_output_file)

    assert {record["PO"] for record in records} == {"PO-TODAY", "PO-YESTERDAY"}
    assert "OUTSIDE_LOOKBACK_WINDOW" in issue_types


def test_invalid_dispatch_dates_are_rejected_and_logged(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={
            RK_SHEET: [
                _row(po="PO-VALID"),
                _row(po="PO-BAD-DATE", dispatch_date="02-04-2025 & 03-04-2025"),
            ]
        },
        include_sheets=(RK_SHEET,),
    )

    result = B2BDispatchPipeline(_config(tmp_path, input_dir, allow_missing=True)).run()

    assert len(_master_records(result.backend_output_file)) == 1
    assert "INVALID_DISPATCH_DATE" in _issue_types(result.backend_output_file)


def test_missing_required_target_sheet_fails_by_default(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={RK_SHEET: [_row()]},
        include_sheets=(RK_SHEET,),
    )

    with pytest.raises(CatastrophicPipelineError):
        B2BDispatchPipeline(_config(tmp_path, input_dir)).run()


def test_missing_required_target_sheet_can_be_allowed(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={RK_SHEET: [_row()]},
        include_sheets=(RK_SHEET,),
    )

    result = B2BDispatchPipeline(_config(tmp_path, input_dir, allow_missing=True)).run()

    assert result.rows_written == 1
    assert "MISSING_TARGET_SHEET" in _issue_types(result.backend_output_file)


def test_missing_critical_headers_fail_the_run(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    headers = [header for header in RK_HEADERS if header != "ASIN"]
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={RK_SHEET: [_row()]},
        include_sheets=(RK_SHEET,),
        header_overrides={RK_SHEET: headers},
    )

    with pytest.raises(CatastrophicPipelineError) as exc:
        B2BDispatchPipeline(_config(tmp_path, input_dir, allow_missing=True)).run()

    assert "Critical headers missing" in str(exc.value)


def test_missing_po_rejects_row(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row(po=""), _row(po="PO-VALID")])

    assert {record["PO"] for record in _master_records(result.backend_output_file)} == {"PO-VALID"}
    assert "MISSING_PO" in _issue_types(result.backend_output_file)


def test_missing_asin_rejects_row(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row(asin=""), _row(asin="ASIN-VALID")])

    assert {record["ASIN"] for record in _master_records(result.backend_output_file)} == {"ASIN-VALID"}
    assert "MISSING_ASIN" in _issue_types(result.backend_output_file)


def test_invalid_dispatch_qty_rejects_row(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row(po="PO-BAD", dispatch_qty="abc"), _row(po="PO-VALID")])

    assert {record["PO"] for record in _master_records(result.backend_output_file)} == {"PO-VALID"}
    assert "INVALID_DISPATCH_QTY" in _issue_types(result.backend_output_file)


def test_po_asin_key_is_derived_if_blank(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row(po="PO-DERIVE", asin="ASIN-DERIVE", po_asin_key="")])
    record = _master_records(result.backend_output_file)[0]

    assert record["PO ASIN Key"] == "PO-DERIVEASIN-DERIVE"
    assert "PO_ASIN_KEY_DERIVED" in _issue_types(result.backend_output_file)


def test_po_asin_key_mismatch_creates_warning(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row(po="PO1", asin="ASIN1", po_asin_key="WRONG")])

    assert "PO_ASIN_KEY_MISMATCH" in _issue_types(result.backend_output_file)
    assert _master_records(result.backend_output_file)[0]["PO ASIN Key"] == "WRONG"


def test_dispatch_value_mismatch_creates_warning(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row(dispatch_qty=3, unit_value=100, dispatch_value_source=250)])

    records = _master_records(result.backend_output_file)
    assert records[0]["Dispatch Value Derived"] == 300
    assert records[0]["Dispatch Value Difference"] == -50
    assert "DISPATCH_VALUE_MISMATCH" in _issue_types(result.backend_output_file)


def test_duplicate_rows_are_audited_but_kept_by_default(tmp_path: Path) -> None:
    duplicate = _row(po="PO-DUP", asin="ASIN-DUP", invoice_no="INV-DUP")
    result = _run_one_sheet(tmp_path, [duplicate, duplicate])

    assert len(_master_records(result.backend_output_file)) == 2
    duplicate_actions = _duplicate_actions(result.backend_output_file)
    assert duplicate_actions == ["KEPT", "KEPT"]


def test_dedupe_exact_rows_drops_later_duplicates_and_audits_them(tmp_path: Path) -> None:
    duplicate = _row(po="PO-DUP", asin="ASIN-DUP", invoice_no="INV-DUP")
    result = _run_one_sheet(tmp_path, [duplicate, duplicate], dedupe=True)

    assert len(_master_records(result.backend_output_file)) == 1
    assert "DROPPED_BY_DEDUPE" in _duplicate_actions(result.backend_output_file)


def test_output_workbook_has_all_required_sheets(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row()])

    wb = load_workbook(result.backend_output_file, data_only=True)
    assert wb.sheetnames == [
        "B2B_Dispatch_Master",
        "Run_Summary",
        "Sheet_Audit",
        "Validation_Issues",
        "Duplicates",
    ]
    wb.close()


def test_latest_copy_is_written(tmp_path: Path) -> None:
    result = _run_one_sheet(tmp_path, [_row()])

    assert result.backend_latest_file.exists()
    assert result.backend_latest_file.name == "B2B_Dispatch_Backend_Audit_latest.xlsx"


def test_reader_reports_clickteck_as_clicktech_channel(tmp_path: Path) -> None:
    workbook_path = tmp_path / "B2B DISPATCH TRACKER.xlsx"
    _write_dispatch_workbook(
        workbook_path,
        rows_by_sheet={CLICK_SHEET: [_row()]},
        include_sheets=(CLICK_SHEET,),
    )

    read_result = read_b2b_dispatch_workbook(workbook_path, _config(tmp_path, tmp_path, allow_missing=True), "RUN1")

    assert read_result.rows[0].source_channel == "CLICKTECH"


def _run_one_sheet(
    tmp_path: Path,
    rows: list[dict[str, object]],
    dedupe: bool = False,
):
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_dispatch_workbook(
        input_dir / "B2B DISPATCH TRACKER.xlsx",
        rows_by_sheet={RK_SHEET: rows},
        include_sheets=(RK_SHEET,),
    )
    return B2BDispatchPipeline(
        _config(tmp_path, input_dir, allow_missing=True, dedupe=dedupe)
    ).run()


def _config(
    tmp_path: Path,
    input_dir: Path,
    allow_missing: bool = False,
    dedupe: bool = False,
) -> B2BDispatchPipelineConfig:
    return B2BDispatchPipelineConfig(
        project_root=tmp_path,
        input_dir=input_dir,
        run_root=tmp_path / "runs",
        processed_dir=tmp_path / "processed" / "b2b_dispatch",
        as_of_date=AS_OF_DATE,
        lookback_days=2,
        allow_missing_target_sheets=allow_missing,
        dedupe_exact_rows=dedupe,
    )


def _row(
    po: str = "PO-1",
    asin: str = "ASIN1",
    invoice_no: str = "INV1",
    dispatch_date: object = "26-06-2026",
    dispatch_qty: object = 5,
    unit_value: object = 100,
    dispatch_value_source: object = 500,
    po_asin_key: str | None = None,
) -> dict[str, object]:
    active_key = f"{po}{asin}" if po_asin_key is None else po_asin_key
    return {
        "Appointment ID": "APT1",
        "Invoice No": invoice_no,
        "Boxes": 2,
        "PO": po,
        "Ship To Location": "BLR1",
        "ASIN": asin,
        "PO ASIN Key": active_key,
        "Model Number": "MODEL1",
        "PO Date": "25-06-2026",
        "PO Qty": 10,
        "Dispatch Qty": dispatch_qty,
        "Unit Value": unit_value,
        "Dispatch Value Source": dispatch_value_source,
        "Dispatch Date": dispatch_date,
        "Dispatch Location": "WH1",
    }


def _write_dispatch_workbook(
    path: Path,
    rows_by_sheet: dict[str, list[dict[str, object]]],
    include_sheets: tuple[str, ...] = (RK_SHEET, CLICK_SHEET, ETRADE_SHEET),
    include_non_target: bool = False,
    header_rows: dict[str, int] | None = None,
    header_overrides: dict[str, list[str]] | None = None,
) -> None:
    wb = Workbook()
    first = True
    for sheet_name in include_sheets:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        first = False
        ws.title = sheet_name
        header_row = (header_rows or {}).get(sheet_name, 3 if sheet_name != ETRADE_SHEET else 4)
        for _ in range(header_row - 1):
            ws.append(["summary/formula row"])
        headers = (header_overrides or {}).get(sheet_name, _headers_for_sheet(sheet_name))
        ws.append(headers)
        for row in rows_by_sheet.get(sheet_name, []):
            ws.append([_value_for_source_header(header, row) for header in headers])

    if include_non_target:
        ws = wb.create_sheet("Blinkit Dispatch")
        ws.append(["summary"])
        ws.append(["summary"])
        ws.append(RK_HEADERS)
        ws.append([_value_for_source_header(header, _row(po="PO-BLINKIT")) for header in RK_HEADERS])

    if first:
        wb.active.title = "Blinkit Dispatch"
    wb.save(path)


def _headers_for_sheet(sheet_name: str) -> list[str]:
    if sheet_name == RK_SHEET:
        return RK_HEADERS
    if sheet_name == CLICK_SHEET:
        return CLICK_HEADERS
    if sheet_name == ETRADE_SHEET:
        return ETRADE_HEADERS
    return RK_HEADERS


def _value_for_source_header(header: str, row: dict[str, object]) -> object:
    source_to_canonical = {
        "APPOINTMENT ID": "Appointment ID",
        "Appointment ID": "Appointment ID",
        "INVOICE NO": "Invoice No",
        "BOXES": "Boxes",
        "PO": "PO",
        "LOC.": "Ship To Location",
        "Ship to location": "Ship To Location",
        "ASIN": "ASIN",
        "PO+ASIN": "PO ASIN Key",
        "SKU": "Model Number",
        "MODEL NAME": "Model Number",
        "PO DATE": "PO Date",
        "PO QTY": "PO Qty",
        "Dispatch Qty": "Dispatch Qty",
        "UNIT VALUE": "Unit Value",
        "UNIT PRICE": "Unit Value",
        "TOTAL VALUE": "Dispatch Value Source",
        "PO DISPATCH VALUE": "Dispatch Value Source",
        "DATE": "Dispatch Date",
        "Date": "Dispatch Date",
        "LOCATION": "Dispatch Location",
        "Location": "Dispatch Location",
    }
    if header == "":
        return None
    return row.get(source_to_canonical[header])


def _master_records(path: Path) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["B2B_Dispatch_Master"]
    headers = [cell.value for cell in ws[1]]
    records = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return records


def _issue_types(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Validation_Issues"]
    headers = [cell.value for cell in ws[1]]
    issue_index = headers.index("Issue Type")
    values = [row[issue_index] for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return values


def _duplicate_actions(path: Path) -> list[str]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Duplicates"]
    headers = [cell.value for cell in ws[1]]
    action_index = headers.index("Action Taken")
    values = [row[action_index] for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return values
