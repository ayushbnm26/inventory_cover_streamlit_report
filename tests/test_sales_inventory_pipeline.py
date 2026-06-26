from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from inventory_cover.config import SalesInventoryPipelineConfig
from inventory_cover.exceptions import CatastrophicPipelineError, FileValidationError
from inventory_cover.io.sales_inventory_excel_io import (
    detect_sales_inventory_header_row_in_sheet,
    read_sales_inventory_workbook,
)
from inventory_cover.pipelines.sales_inventory_pipeline import SalesInventoryPipeline
from inventory_cover.sales_inventory_schemas import INVENTORY_REPORT_TYPE, SALES_REPORT_TYPE
from inventory_cover.cli import main as cli_main


SALES_HEADERS = [
    "ASIN",
    "Child Vendor Code",
    "Product Title",
    "Brand Code",
    "Brand",
    "Category",
    "Subcategory",
    "Parent ASIN",
    "UPC",
    "EAN",
    "ISBN",
    "Model Number",
    "Store Code",
    "MSRP",
    "Binding",
    "Colour",
    "Release Date",
    "Replenishment Code",
    "Shipped Revenue",
    "Shipped COGS",
    "Shipped Units",
    "Customer Returns",
    "Confirmed Units",
    "Sales Discount",
    "Contra-COGS",
    "Net PPM %",
    "ASIN Confirmation %",
]

INVENTORY_HEADERS = [
    "ASIN",
    "Child Vendor Code",
    "Product Title",
    "Brand Code",
    "Brand",
    "Category",
    "Subcategory",
    "Parent ASIN",
    "UPC",
    "EAN",
    "ISBN",
    "Model Number",
    "Store Code",
    "MSRP",
    "Binding",
    "Colour",
    "Release Date",
    "Replenishment Code",
    "Vendor Confirmation %",
    "Net Received",
    "Net Received Units",
    "Open Purchase Order Quantity",
    "Receive Fill %",
    "Overall Vendor Lead Time (days)",
    "Aged 90+ Days Sellable Inventory",
    "Aged 90+ Days Sellable Units",
    "Sellable On-Hand Inventory",
    "Sellable On Hand Units",
    "Unsellable On-Hand Inventory",
    "Unsellable On-Hand Units",
    "Confirmed Units",
    "Sales Discount",
    "Contra-COGS",
    "In Transit Quantity",
    "Sellable In Transit Units",
    "Unsellable In Transit Units",
]


def test_sales_header_detection_works_when_headers_are_on_row_2() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append([_metadata_text("01/05/26 - 31/05/26")])
    ws.append(SALES_HEADERS)

    detection = detect_sales_inventory_header_row_in_sheet(ws, SALES_REPORT_TYPE)

    assert detection is not None
    assert detection.header_row == 2
    assert detection.missing_minimum_headers == ()


def test_inventory_header_detection_works_when_headers_are_on_row_2() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    ws.append([_metadata_text("24/06/26 - 24/06/26")])
    ws.append(INVENTORY_HEADERS)

    detection = detect_sales_inventory_header_row_in_sheet(ws, INVENTORY_REPORT_TYPE)

    assert detection is not None
    assert detection.header_row == 2
    assert detection.missing_minimum_headers == ()


def test_header_detection_works_when_metadata_rows_shift() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["metadata"])
    ws.append(["more metadata"])
    ws.append(["another metadata row"])
    ws.append(SALES_HEADERS)

    detection = detect_sales_inventory_header_row_in_sheet(ws, SALES_REPORT_TYPE)

    assert detection is not None
    assert detection.header_row == 4


def test_sales_file_processes_observed_expected_columns(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row(asin="ASIN-SALES")], inventory_rows=[])
    records = _records(result.sales_output_file, "Sales_Master")

    assert result.sales_rows_written == 1
    assert records[0]["Report Type"] == "SALES"
    assert records[0]["ASIN"] == "ASIN-SALES"
    assert records[0]["Currency"] == "INR"
    assert records[0]["Viewing Range Start"] == datetime(2026, 5, 1)


def test_inventory_file_processes_observed_expected_columns(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[], inventory_rows=[_inventory_row(asin="ASIN-INV")])
    records = _records(result.inventory_output_file, "Inventory_Master")

    assert result.inventory_rows_written == 1
    assert records[0]["Report Type"] == "INVENTORY"
    assert records[0]["ASIN"] == "ASIN-INV"
    assert records[0]["Sellable On Hand Units"] == 8


def test_extra_unknown_columns_do_not_break_the_run(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[_sales_row()],
        inventory_rows=[],
        sales_headers=SALES_HEADERS + ["Unexpected Column"],
    )

    assert result.sales_rows_written == 1
    assert "EXTRA_SOURCE_COLUMNS" in _issue_types(result.sales_output_file)


def test_missing_optional_sales_columns_create_warnings_but_do_not_fail(tmp_path: Path) -> None:
    headers = ["ASIN", "Child Vendor Code", "Product Title", "Model Number", "Shipped Units"]
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row()], inventory_rows=[], sales_headers=headers)

    assert result.sales_rows_written == 1
    assert "MISSING_EXPECTED_COLUMNS" in _issue_types(result.sales_output_file)


def test_missing_optional_inventory_columns_create_warnings_but_do_not_fail(tmp_path: Path) -> None:
    headers = ["ASIN", "Child Vendor Code", "Product Title", "Model Number", "Sellable On Hand Units"]
    result = _run_pipeline(tmp_path, sales_rows=[], inventory_rows=[_inventory_row()], inventory_headers=headers)

    assert result.inventory_rows_written == 1
    assert "MISSING_EXPECTED_COLUMNS" in _issue_types(result.inventory_output_file)


def test_missing_minimum_identifying_headers_causes_file_level_failure(tmp_path: Path) -> None:
    workbook_path = tmp_path / "bad-sales.xlsx"
    headers = [header for header in SALES_HEADERS if header != "Shipped Units"]
    _write_vendor_workbook(workbook_path, SALES_REPORT_TYPE, [_sales_row()], headers=headers)

    with pytest.raises(FileValidationError):
        read_sales_inventory_workbook(workbook_path, _config(tmp_path), SALES_REPORT_TYPE)


def test_blank_rows_are_skipped(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[{}, _sales_row()], inventory_rows=[])
    records = _records(result.sales_output_file, "Sales_Master")
    audit = _records(result.sales_output_file, "File_Audit")

    assert len(records) == 1
    assert audit[0]["Rows Blank Skipped"] == 1


def test_rows_with_asin_are_accepted(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row(asin="ASIN-OK")], inventory_rows=[])

    assert result.sales_rows_written == 1


def test_rows_without_asin_but_with_model_and_title_are_kept_with_warning(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row(asin="", model_number="MODEL-X")], inventory_rows=[])
    records = _records(result.sales_output_file, "Sales_Master")

    assert len(records) == 1
    assert records[0]["Row Validation Status"] == "WARNING"
    assert "MISSING_ASIN_IDENTIFIER" in _issue_types(result.sales_output_file)


def test_blank_model_number_is_filled_from_asin_mapping(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[_sales_row(asin="ASIN-MAP", model_number="")],
        inventory_rows=[],
        mapping_rows=[{"ASIN": "ASIN-MAP", "SKU": "SKU-MAPPED", "Master SKU": "SKU-MAPPED"}],
    )
    record = _records(result.sales_output_file, "Sales_Master")[0]

    assert record["Model Number"] == "SKU-MAPPED"
    assert "MODEL_NUMBER_FILLED_FROM_MAPPING" in _issue_types(result.sales_output_file)
    assert "MISSING_MODEL_NUMBER" not in _issue_types(result.sales_output_file)


def test_blank_asin_is_filled_from_model_mapping(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[_sales_row(asin="", model_number="SKU-MAPPED")],
        inventory_rows=[],
        mapping_rows=[{"ASIN": "ASIN-MAP", "SKU": "SKU-MAPPED", "Master SKU": "SKU-MAPPED"}],
    )
    record = _records(result.sales_output_file, "Sales_Master")[0]

    assert record["ASIN"] == "ASIN-MAP"
    assert "ASIN_FILLED_FROM_MAPPING" in _issue_types(result.sales_output_file)
    assert "MISSING_ASIN_IDENTIFIER" not in _issue_types(result.sales_output_file)


def test_completely_unusable_rows_are_rejected(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[
            _sales_row(asin="", model_number="", product_title=""),
            _sales_row(asin="ASIN-VALID"),
        ],
        inventory_rows=[],
    )
    records = _records(result.sales_output_file, "Sales_Master")

    assert len(records) == 1
    assert "UNUSABLE_ROW" in _issue_types(result.sales_output_file)


def test_numeric_parsing_handles_commas_and_blanks(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[_sales_row(shipped_revenue="1,234.50", customer_returns="")],
        inventory_rows=[],
    )
    record = _records(result.sales_output_file, "Sales_Master")[0]

    assert record["Shipped Revenue"] == 1234.5
    assert record["Customer Returns"] is None


def test_date_parsing_handles_excel_dates_and_text_dates(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[_sales_row(release_date=date(2026, 6, 24))],
        inventory_rows=[_inventory_row(release_date="24-06-2026")],
    )

    assert _records(result.sales_output_file, "Sales_Master")[0]["Release Date"] == datetime(2026, 6, 24)
    assert _records(result.inventory_output_file, "Inventory_Master")[0]["Release Date"] == datetime(2026, 6, 24)


def test_invalid_release_date_creates_warning_not_rejection(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row(release_date="not-a-date")], inventory_rows=[])
    record = _records(result.sales_output_file, "Sales_Master")[0]

    assert record["Release Date"] == "not-a-date"
    assert "INVALID_RELEASE_DATE" in _issue_types(result.sales_output_file)


def test_duplicates_are_audited_but_kept_by_default(tmp_path: Path) -> None:
    duplicate = _sales_row(asin="ASIN-DUP")
    result = _run_pipeline(tmp_path, sales_rows=[duplicate, duplicate], inventory_rows=[])

    assert len(_records(result.sales_output_file, "Sales_Master")) == 2
    assert _duplicate_actions(result.sales_output_file) == ["KEPT", "KEPT"]


def test_exact_duplicates_are_dropped_only_with_dedupe_flag(tmp_path: Path) -> None:
    duplicate = _sales_row(asin="ASIN-DUP")
    result = _run_pipeline(tmp_path, sales_rows=[duplicate, duplicate], inventory_rows=[], dedupe=True)

    assert len(_records(result.sales_output_file, "Sales_Master")) == 1
    assert "DROPPED_BY_DEDUPE" in _duplicate_actions(result.sales_output_file)


def test_sales_only_run_works(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row()], inventory_rows=[])

    assert result.sales_output_file is not None
    assert result.inventory_output_file is None


def test_inventory_only_run_works(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[], inventory_rows=[_inventory_row()])

    assert result.sales_output_file is None
    assert result.inventory_output_file is not None


def test_blank_inventory_stock_and_intransit_quantities_default_to_zero(tmp_path: Path) -> None:
    result = _run_pipeline(
        tmp_path,
        sales_rows=[],
        inventory_rows=[
            _inventory_row(
                sellable_on_hand_inventory="",
                sellable_on_hand_units="",
                in_transit_quantity="",
                sellable_in_transit_units="",
                unsellable_in_transit_units="",
            )
        ],
    )
    record = _records(result.inventory_output_file, "Inventory_Master")[0]
    issue_types = _issue_types(result.inventory_output_file)

    assert record["Sellable On-Hand Inventory"] == 0
    assert record["Sellable On Hand Units"] == 0
    assert record["In Transit Quantity"] == 0
    assert record["Sellable In Transit Units"] == 0
    assert record["Unsellable In Transit Units"] == 0
    assert "BLANK_NUMERIC_DEFAULTED_TO_ZERO" not in issue_types
    assert "MISSING_SELLABLE_ON_HAND_UNITS" not in issue_types
    assert "MISSING_IN_TRANSIT_QUANTITY" not in issue_types


def test_both_missing_run_fails(tmp_path: Path) -> None:
    with pytest.raises(CatastrophicPipelineError):
        SalesInventoryPipeline(_config(tmp_path)).run()


def test_multiple_files_fail_by_default(tmp_path: Path) -> None:
    sales_dir = tmp_path / "incoming" / "sales"
    sales_dir.mkdir(parents=True)
    _write_vendor_workbook(sales_dir / "sales1.xlsx", SALES_REPORT_TYPE, [_sales_row()])
    _write_vendor_workbook(sales_dir / "sales2.xlsx", SALES_REPORT_TYPE, [_sales_row(asin="ASIN2")])

    with pytest.raises(CatastrophicPipelineError):
        SalesInventoryPipeline(_config(tmp_path)).run()


def test_multiple_files_process_only_with_allow_flags(tmp_path: Path) -> None:
    sales_dir = tmp_path / "incoming" / "sales"
    sales_dir.mkdir(parents=True)
    _write_vendor_workbook(sales_dir / "sales1.xlsx", SALES_REPORT_TYPE, [_sales_row(asin="ASIN1")])
    _write_vendor_workbook(sales_dir / "sales2.xlsx", SALES_REPORT_TYPE, [_sales_row(asin="ASIN2")])

    result = SalesInventoryPipeline(_config(tmp_path, allow_multiple_sales=True)).run()

    assert result.sales_rows_written == 2


def test_output_workbooks_contain_required_sheets_and_processing_guides(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row()], inventory_rows=[_inventory_row()])

    sales_wb = load_workbook(result.sales_output_file, data_only=True)
    inventory_wb = load_workbook(result.inventory_output_file, data_only=True)
    assert sales_wb.sheetnames == [
        "Sales_Master",
        "Run_Summary",
        "File_Audit",
        "Validation_Issues",
        "Duplicates",
        "Mapping_Audit",
        "Processing_Guide",
    ]
    assert inventory_wb.sheetnames == [
        "Inventory_Master",
        "Run_Summary",
        "File_Audit",
        "Validation_Issues",
        "Duplicates",
        "Mapping_Audit",
        "Processing_Guide",
    ]
    assert sales_wb["Processing_Guide"].max_row > 10
    assert inventory_wb["Processing_Guide"].max_row > 10
    sales_wb.close()
    inventory_wb.close()


def test_latest_copies_are_written(tmp_path: Path) -> None:
    result = _run_pipeline(tmp_path, sales_rows=[_sales_row()], inventory_rows=[_inventory_row()])

    assert result.latest_sales_backend_file.exists()
    assert result.latest_inventory_backend_file.exists()
    assert result.latest_run_summary_file.exists()


def test_cli_command_run_sales_inventory_works(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    sales_dir = tmp_path / "incoming" / "sales"
    sales_dir.mkdir(parents=True)
    inventory_dir = tmp_path / "incoming" / "inventory"
    inventory_dir.mkdir(parents=True)
    _write_vendor_workbook(sales_dir / "sales.xlsx", SALES_REPORT_TYPE, [_sales_row()])

    code = cli_main(
        [
            "run-sales-inventory",
            "--sales-input-dir",
            str(sales_dir),
            "--inventory-input-dir",
            str(inventory_dir),
            "--mapping-input-dir",
            str(tmp_path / "reference" / "sales_inventory_mapping"),
            "--run-root",
            str(tmp_path / "runs"),
            "--processed-dir",
            str(tmp_path / "processed" / "sales_inventory"),
        ]
    )

    output = capsys.readouterr().out
    assert code == 0
    assert "Sales backend workbook:" in output
    assert "Inventory backend workbook: NOT GENERATED" in output


def test_centralized_cli_can_run_selected_source_pipeline(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    sales_dir = tmp_path / "incoming" / "sales"
    sales_dir.mkdir(parents=True)
    inventory_dir = tmp_path / "incoming" / "inventory"
    inventory_dir.mkdir(parents=True)
    mapping_dir = tmp_path / "reference" / "sales_inventory_mapping"
    mapping_dir.mkdir(parents=True)
    _write_vendor_workbook(sales_dir / "sales.xlsx", SALES_REPORT_TYPE, [_sales_row()])

    code = cli_main(
        [
            "run-source-pipelines",
            "--skip-po-items",
            "--skip-b2b-dispatch",
            "--sales-input-dir",
            str(sales_dir),
            "--inventory-input-dir",
            str(inventory_dir),
            "--mapping-input-dir",
            str(mapping_dir),
            "--run-root",
            str(tmp_path / "runs"),
            "--sales-inventory-processed-dir",
            str(tmp_path / "processed" / "sales_inventory"),
            "--min-free-gb",
            "0",
        ]
    )

    output = capsys.readouterr().out
    assert code == 0
    assert "Central source pipeline run mode: sequential" in output
    assert "Pipeline 3: Sales & Inventory" in output
    assert "Rows written: sales=1, inventory=0" in output


def _run_pipeline(
    tmp_path: Path,
    sales_rows: list[dict[str, object]],
    inventory_rows: list[dict[str, object]],
    sales_headers: list[str] | None = None,
    inventory_headers: list[str] | None = None,
    mapping_rows: list[dict[str, object]] | None = None,
    dedupe: bool = False,
):
    sales_dir = tmp_path / "incoming" / "sales"
    inventory_dir = tmp_path / "incoming" / "inventory"
    sales_dir.mkdir(parents=True)
    inventory_dir.mkdir(parents=True)
    if sales_rows:
        _write_vendor_workbook(
            sales_dir / "sales.xlsx",
            SALES_REPORT_TYPE,
            sales_rows,
            headers=sales_headers or SALES_HEADERS,
        )
    if inventory_rows:
        _write_vendor_workbook(
            inventory_dir / "inventory.xlsx",
            INVENTORY_REPORT_TYPE,
            inventory_rows,
            headers=inventory_headers or INVENTORY_HEADERS,
        )
    mapping_dir = tmp_path / "reference" / "sales_inventory_mapping"
    if mapping_rows is not None:
        mapping_dir.mkdir(parents=True)
        _write_mapping_workbook(mapping_dir / "mapping.xlsx", mapping_rows)
    return SalesInventoryPipeline(_config(tmp_path, dedupe=dedupe)).run()


def _config(
    tmp_path: Path,
    dedupe: bool = False,
    allow_multiple_sales: bool = False,
) -> SalesInventoryPipelineConfig:
    return SalesInventoryPipelineConfig(
        project_root=tmp_path,
        sales_input_dir=tmp_path / "incoming" / "sales",
        inventory_input_dir=tmp_path / "incoming" / "inventory",
        mapping_input_dir=tmp_path / "reference" / "sales_inventory_mapping",
        run_root=tmp_path / "runs",
        processed_dir=tmp_path / "processed" / "sales_inventory",
        dedupe_exact_rows=dedupe,
        allow_multiple_sales_files=allow_multiple_sales,
    )


def _write_vendor_workbook(
    path: Path,
    report_type: str,
    rows: list[dict[str, object]],
    headers: list[str] | None = None,
    header_row: int = 2,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet0"
    for row_number in range(1, header_row):
        if row_number == 1:
            ws.append([_metadata_text("01/05/26 - 31/05/26" if report_type == SALES_REPORT_TYPE else "24/06/26 - 24/06/26")])
        else:
            ws.append(["metadata"])
    active_headers = headers or (SALES_HEADERS if report_type == SALES_REPORT_TYPE else INVENTORY_HEADERS)
    ws.append(active_headers)
    for row in rows:
        ws.append([_value_for_header(header, row) for header in active_headers])
    wb.save(path)


def _metadata_text(viewing_range: str) -> str:
    return (
        "Programme=[Retail] Distributor View=[Sourcing] View By=[ASIN] Countries=[IN] "
        "Businesses=[Test Business] Locale=[en_IN] Currency=[INR] Reporting Range=[Custom] "
        f"Viewing Range=[{viewing_range}] Report Updated=[24/06/26]"
    )


def _sales_row(
    asin: str = "ASIN1",
    model_number: str = "MODEL1",
    product_title: str = "Synthetic Sales Product",
    release_date: object = "24/06/26",
    shipped_revenue: object = 1000,
    customer_returns: object = 0,
) -> dict[str, object]:
    return {
        "ASIN": asin,
        "Child Vendor Code": "VENDOR1",
        "Product Title": product_title,
        "Brand Code": "BR1",
        "Brand": "Brand",
        "Category": "Category",
        "Subcategory": "Subcategory",
        "Parent ASIN": "PARENT1",
        "UPC": "001234567890",
        "EAN": "0001234567890",
        "ISBN": "0000000000",
        "Model Number": model_number,
        "Store Code": "STORE",
        "MSRP": "1,499",
        "Binding": "Electronics",
        "Colour": "Black",
        "Release Date": release_date,
        "Replenishment Code": "ALLOC",
        "Shipped Revenue": shipped_revenue,
        "Shipped COGS": 700,
        "Shipped Units": 10,
        "Customer Returns": customer_returns,
        "Confirmed Units": 9,
        "Sales Discount": "",
        "Contra-COGS": "",
        "Net PPM %": 12.5,
        "ASIN Confirmation %": 95,
    }


def _inventory_row(
    asin: str = "ASIN1",
    release_date: object = "24/06/26",
    sellable_on_hand_inventory: object = 2000,
    sellable_on_hand_units: object = 8,
    in_transit_quantity: object = 2,
    sellable_in_transit_units: object = 2,
    unsellable_in_transit_units: object = 0,
) -> dict[str, object]:
    base = _sales_row(asin=asin, release_date=release_date)
    return {
        **base,
        "Vendor Confirmation %": 90,
        "Net Received": 1000,
        "Net Received Units": 5,
        "Open Purchase Order Quantity": 3,
        "Receive Fill %": 80,
        "Overall Vendor Lead Time (days)": 7,
        "Aged 90+ Days Sellable Inventory": 0,
        "Aged 90+ Days Sellable Units": 0,
        "Sellable On-Hand Inventory": sellable_on_hand_inventory,
        "Sellable On Hand Units": sellable_on_hand_units,
        "Unsellable On-Hand Inventory": "",
        "Unsellable On-Hand Units": "",
        "In Transit Quantity": in_transit_quantity,
        "Sellable In Transit Units": sellable_in_transit_units,
        "Unsellable In Transit Units": unsellable_in_transit_units,
    }


def _write_mapping_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["QUICK COM"])
    headers = ["ASIN", "SKU", "Master SKU", "Brand"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    wb.save(path)


def _value_for_header(header: str, row: dict[str, object]) -> object:
    return row.get(header)


def _records(path: Path | None, sheet_name: str) -> list[dict[str, object]]:
    assert path is not None
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    records = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return records


def _issue_types(path: Path | None) -> list[str]:
    records = _records(path, "Validation_Issues")
    return [str(record["Issue Type"]) for record in records]


def _duplicate_actions(path: Path | None) -> list[str]:
    records = _records(path, "Duplicates")
    return [str(record["Action Taken"]) for record in records]
