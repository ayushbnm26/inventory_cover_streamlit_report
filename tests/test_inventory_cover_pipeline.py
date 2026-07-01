from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from inventory_cover.calculations.inventory_cover_builder import classify_bucket, compute_product
from inventory_cover.cli import main as cli_main
from inventory_cover.config import InventoryCoverPipelineConfig
from inventory_cover.exceptions import CatastrophicPipelineError
from inventory_cover.inventory_cover_schemas import ProductCoverRow
from inventory_cover.pipelines.inventory_cover_pipeline import InventoryCoverPipeline


SALES_HEADERS = [
    "Run ID", "Source File", "Source Sheet", "Source Row", "ASIN", "Child Vendor Code",
    "Product Title", "Brand", "Category", "Subcategory", "Model Number",
    "Viewing Range Start", "Viewing Range End", "Report Updated Date", "Shipped Units",
]
INVENTORY_HEADERS = [
    "Run ID", "Source File", "Source Sheet", "Source Row", "ASIN", "Child Vendor Code",
    "Product Title", "Brand", "Category", "Subcategory", "Model Number",
    "Viewing Range Start", "Viewing Range End", "Report Updated Date",
    "Sellable On Hand Units", "Sellable In Transit Units", "In Transit Quantity",
]
B2B_HEADERS = [
    "Run ID", "Source File", "Source Sheet", "Source Row", "ASIN", "Model Number",
    "Dispatch Qty", "Dispatch Date", "Included In Lookback Window",
]
PO_HEADERS = [
    "Run ID", "Source File", "Source Sheet", "Source Row", "ASIN", "Model Number",
    "PO", "Quantity Outstanding", "Open PO Qty - Final", "Window Start", "Window End",
]
ASIN_HEADERS = [
    "ASIN", "SKU", "Brand", "Brand Name", "Vendor", "Main Category", "Sub Category", "Aligned DOH Target",
]


def _write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _read_table(path: Path, sheet_name: str, data_only: bool = True) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=data_only)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        out.append({headers[i]: raw[i] for i in range(len(headers))})
    return out


def _formula_cell(path: Path, sheet_name: str, header: str, row_index: int = 2) -> Any:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name]
    col = {cell.value: cell.column for cell in ws[1]}[header]
    return ws.cell(row=row_index, column=col).value


def make_sources(
    tmp_path: Path,
    *,
    sales_rows: list[dict[str, Any]] | None = None,
    inventory_rows: list[dict[str, Any]] | None = None,
    b2b_rows: list[dict[str, Any]] | None = None,
    po_rows: list[dict[str, Any]] | None = None,
    asin_rows: list[dict[str, Any]] | None = None,
    write_sales: bool = True,
    write_inventory: bool = True,
    write_b2b: bool = True,
    write_po: bool = True,
    write_asin: bool = True,
) -> InventoryCoverPipelineConfig:
    base = tmp_path / "data"
    sales_path = base / "sales.xlsx"
    inventory_path = base / "inventory.xlsx"
    b2b_path = base / "b2b.xlsx"
    po_path = base / "po.xlsx"
    asin_path = base / "asin.xlsx"

    if write_sales:
        _write_workbook(sales_path, "Sales_Master", SALES_HEADERS, sales_rows or [])
    if write_inventory:
        _write_workbook(inventory_path, "Inventory_Master", INVENTORY_HEADERS, inventory_rows or [])
    if write_b2b:
        _write_workbook(b2b_path, "B2B_Dispatch_Master", B2B_HEADERS, b2b_rows or [])
    if write_po:
        _write_workbook(po_path, "PO_Items_Master", PO_HEADERS, po_rows or [])
    if write_asin:
        _write_workbook(asin_path, "ASIN_Master", ASIN_HEADERS, asin_rows or [])

    return InventoryCoverPipelineConfig(
        project_root=tmp_path,
        sales_backend_path=sales_path,
        inventory_backend_path=inventory_path,
        b2b_backend_path=b2b_path,
        po_backend_path=po_path,
        asin_master_path=asin_path,
        run_root=tmp_path / "runs",
        processed_dir=tmp_path / "processed" / "inventory_cover",
    )


def _full_window_sales(asin: str, units: int) -> dict[str, Any]:
    return {
        "ASIN": asin, "Model Number": f"M-{asin}", "Brand": "BR", "Category": "Cat",
        "Subcategory": "Sub", "Child Vendor Code": "V1",
        "Viewing Range Start": date(2024, 5, 1), "Viewing Range End": date(2024, 5, 30),
        "Report Updated Date": date(2024, 5, 31), "Shipped Units": units,
    }


def _inventory_row(asin: str, on_hand: Any, amz: Any = 0, transit: Any = None) -> dict[str, Any]:
    return {
        "ASIN": asin, "Model Number": f"M-{asin}", "Brand": "BR", "Category": "Cat",
        "Subcategory": "Sub", "Child Vendor Code": "V1",
        "Viewing Range End": date(2024, 5, 31), "Report Updated Date": date(2024, 5, 31),
        "Sellable On Hand Units": on_hand, "Sellable In Transit Units": amz, "In Transit Quantity": transit,
    }


# ---------------------------------------------------------------------------
# Pure calculation tests.
# ---------------------------------------------------------------------------
def _product_with_cover(target_value: float) -> ProductCoverRow:
    product = ProductCoverRow(product_key="ASIN::A", product_key_type="ASIN", asin="A", model_number="M")
    product.sales_period_start = date(2024, 5, 1)
    product.sales_period_end = date(2024, 5, 30)
    product.sales_units = 30
    product.on_hand_units = target_value
    return product


class _Cfg:
    sales_window_days = 30
    default_target_doh = 30.0


def test_classify_bucket_boundaries() -> None:
    assert classify_bucket(4) == "Critical"
    assert classify_bucket(5) == "High Risk"
    assert classify_bucket(14.9) == "High Risk"
    assert classify_bucket(15) == "Watch"
    assert classify_bucket(24.9) == "Watch"
    assert classify_bucket(25) == "Near Target"
    assert classify_bucket(30) == "Near Target"
    assert classify_bucket(30.1) == "Healthy"
    assert classify_bucket("No Sales") == "No Sales"


@pytest.mark.parametrize(
    "cover,bucket",
    [(4, "Critical"), (10, "High Risk"), (20, "Watch"), (28, "Near Target"), (40, "Healthy")],
)
def test_compute_product_buckets(cover: float, bucket: str) -> None:
    product = _product_with_cover(cover)
    compute_product(product, _Cfg())
    assert product.daily_run_rate == pytest.approx(1.0)
    assert product.total_supply_cover_doh == pytest.approx(cover)
    assert product.cover_bucket == bucket


def test_sales_days_31_days_capped_to_30() -> None:
    product = _product_with_cover(10)
    product.sales_period_start = date(2024, 5, 1)
    product.sales_period_end = date(2024, 5, 31)  # 31 days inclusive
    compute_product(product, _Cfg())
    assert product.sales_days == 30


def test_sales_days_7_days() -> None:
    product = _product_with_cover(10)
    product.sales_period_start = date(2024, 5, 25)
    product.sales_period_end = date(2024, 5, 31)  # 7 days inclusive
    compute_product(product, _Cfg())
    assert product.sales_days == 7


def test_divide_by_zero_returns_no_sales() -> None:
    product = ProductCoverRow(product_key="ASIN::A", product_key_type="ASIN", asin="A")
    product.on_hand_units = 100
    compute_product(product, _Cfg())
    assert product.daily_run_rate == 0
    assert product.current_stock_doh == "No Sales"
    assert product.total_supply_cover_doh == "No Sales"
    assert product.cover_bucket == "No Sales"


# ---------------------------------------------------------------------------
# Pipeline integration tests.
# ---------------------------------------------------------------------------
def test_engine_runs_with_all_sources(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100, amz=10)],
        b2b_rows=[{"ASIN": "A1", "Model Number": "M-A1", "Dispatch Qty": 5, "Included In Lookback Window": True}],
        po_rows=[{"ASIN": "A1", "Model Number": "M-A1", "PO": "PO1", "Open PO Qty - Final": 20}],
        asin_rows=[{"ASIN": "A1", "SKU": "SKU-A1", "Brand": "BR", "Brand Name": "Brand One",
                    "Vendor": "Vendor X", "Main Category": "MC", "Sub Category": "SC", "Aligned DOH Target": 45}],
    )
    result = InventoryCoverPipeline(config).run()
    assert result.product_count == 1
    assert result.team_output_file.exists()
    assert result.backend_output_file.exists()
    assert result.team_latest_file.exists()
    assert result.backend_latest_file.exists()

    master = _read_table(result.backend_output_file, "Inventory_Cover_Master")
    row = master[0]
    assert row["ASIN"] == "A1"
    assert row["Model Number / SKU"] == "SKU-A1"
    assert row["Brand Name"] == "Brand One"
    assert row["Target DOH"] == 45
    assert row["Sellable On Hand Units"] == 100
    assert row["Amazon In-Transit Units"] == 10
    assert row["Own In-Transit Units"] == 5
    assert row["Open PO Quantity"] == 20


def test_engine_runs_when_asin_master_missing(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100)],
        write_asin=False,
    )
    result = InventoryCoverPipeline(config).run()
    assert result.product_count == 1
    team = _read_table(result.team_output_file, "Inventory_Cover_Report")
    master = _read_table(result.backend_output_file, "Inventory_Cover_Master")
    assert team[0]["Aligned DOH Target"] == 30
    assert master[0]["Target DOH"] == 30  # default target


def test_team_workbook_omits_data_quality_flag_but_backend_keeps_it(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100)],
    )
    result = InventoryCoverPipeline(config).run()
    team = _read_table(result.team_output_file, "Inventory_Cover_Report")
    backend = _read_table(result.backend_output_file, "Inventory_Cover_Master")
    assert "Data Quality Flag" not in team[0]
    assert "Data Quality Flag" in backend[0]


def test_all_sources_missing_raises(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        write_sales=False, write_inventory=False, write_b2b=False, write_po=False, write_asin=False,
    )
    with pytest.raises(CatastrophicPipelineError):
        InventoryCoverPipeline(config).run()


def test_product_universe_includes_inventory_only(tmp_path: Path) -> None:
    config = make_sources(tmp_path, inventory_rows=[_inventory_row("INV1", 50)])
    result = InventoryCoverPipeline(config).run()
    asins = {row["ASIN"] for row in _read_table(result.backend_output_file, "Inventory_Cover_Master")}
    assert "INV1" in asins


def test_product_universe_includes_sales_only(tmp_path: Path) -> None:
    config = make_sources(tmp_path, sales_rows=[_full_window_sales("SAL1", 30)])
    result = InventoryCoverPipeline(config).run()
    asins = {row["ASIN"] for row in _read_table(result.backend_output_file, "Inventory_Cover_Master")}
    assert "SAL1" in asins


def test_product_universe_includes_po_only(tmp_path: Path) -> None:
    config = make_sources(tmp_path, po_rows=[{"ASIN": "PO1", "Model Number": "M", "Open PO Qty - Final": 5}])
    result = InventoryCoverPipeline(config).run()
    asins = {row["ASIN"] for row in _read_table(result.backend_output_file, "Inventory_Cover_Master")}
    assert "PO1" in asins


def test_product_universe_includes_b2b_only(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        b2b_rows=[{"ASIN": "B2B1", "Model Number": "M", "Dispatch Qty": 9, "Included In Lookback Window": True}],
    )
    result = InventoryCoverPipeline(config).run()
    rows = _read_table(result.backend_output_file, "Inventory_Cover_Master")
    match = [r for r in rows if r["ASIN"] == "B2B1"]
    assert match and match[0]["Own In-Transit Units"] == 9


def test_blank_numeric_treated_as_zero(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", None, amz=None, transit=None)],
    )
    result = InventoryCoverPipeline(config).run()
    row = _read_table(result.backend_output_file, "Inventory_Cover_Master")[0]
    assert row["Sellable On Hand Units"] == 0
    assert row["Amazon In-Transit Units"] == 0


def test_raw_source_not_overwritten(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100)],
    )
    before = _read_table(config.inventory_backend_path, "Inventory_Master")
    InventoryCoverPipeline(config).run()
    after = _read_table(config.inventory_backend_path, "Inventory_Master")
    assert before == after


def test_team_cells_have_formulas_with_cached_values(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100, amz=10)],
    )
    result = InventoryCoverPipeline(config).run()
    rows = _read_table(result.team_output_file, "Inventory_Cover_Report")
    assert rows[0]["Daily Run Rate"] == pytest.approx(1.0)
    assert rows[0]["Current Stock DOH"] == pytest.approx(100.0)

    drr = _formula_cell(result.team_output_file, "Inventory_Cover_Report", "Daily Run Rate")
    assert isinstance(drr, str) and drr.startswith("=") and "K2" in drr
    assert "[@" not in drr
    for header in ("Current Stock DOH", "Total Supply Cover DOH", "Cover Bucket", "Gap to Target Units"):
        formula = _formula_cell(result.team_output_file, "Inventory_Cover_Report", header)
        assert isinstance(formula, str) and formula.startswith("=")

    audit_drr = _formula_cell(result.team_output_file, "Formula_Audit", "Daily Run Rate")
    assert isinstance(audit_drr, str) and audit_drr.startswith("=")


def test_team_workbook_has_no_excel_table_parts(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100, amz=10)],
    )
    result = InventoryCoverPipeline(config).run()
    with ZipFile(result.team_output_file) as workbook:
        assert not any(name.startswith("xl/tables/") for name in workbook.namelist())


def test_no_sales_product_in_annexure(tmp_path: Path) -> None:
    config = make_sources(tmp_path, inventory_rows=[_inventory_row("NS1", 100)])
    result = InventoryCoverPipeline(config).run()
    no_sales = _read_table(result.team_output_file, "No_Sales")
    asins = {row.get("ASIN") for row in no_sales}
    assert "NS1" in asins


def test_critical_product_in_annexure(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("C1", 30)],
        inventory_rows=[_inventory_row("C1", 3)],
    )
    result = InventoryCoverPipeline(config).run()
    critical = _read_table(result.team_output_file, "Critical")
    assert {row.get("ASIN") for row in critical} == {"C1"}


def test_team_and_backend_sheets_present(tmp_path: Path) -> None:
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100)],
    )
    result = InventoryCoverPipeline(config).run()
    team = load_workbook(result.team_output_file).sheetnames
    backend = load_workbook(result.backend_output_file).sheetnames
    for sheet in ("Inventory_Cover_Report", "Critical", "High_Risk", "Watch", "Near_Target",
                  "Healthy", "No_Sales", "Formula_Guide", "Source_Summary", "Formula_Audit"):
        assert sheet in team
    for sheet in ("Inventory_Cover_Master", "Source_Row_Trace", "Source_Summary",
                  "Validation_Issues", "Calculation_Audit", "Formula_Guide", "Run_Metadata"):
        assert sheet in backend


def test_cli_run_inventory_cover(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("GDRIVE_ENABLED", "false")
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100)],
    )
    rc = cli_main([
        "run-inventory-cover",
        "--sales-backend-path", str(config.sales_backend_path),
        "--inventory-backend-path", str(config.inventory_backend_path),
        "--b2b-backend-path", str(config.b2b_backend_path),
        "--po-backend-path", str(config.po_backend_path),
        "--asin-master-path", str(config.asin_master_path),
        "--run-root", str(config.run_root),
        "--processed-dir", str(config.processed_dir),
    ])
    assert rc == 0


def test_cli_run_full_inventory_cover_skip_sources(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("GDRIVE_ENABLED", "false")
    config = make_sources(
        tmp_path,
        sales_rows=[_full_window_sales("A1", 30)],
        inventory_rows=[_inventory_row("A1", 100)],
    )
    rc = cli_main([
        "run-full-inventory-cover",
        "--skip-source-pipelines",
        "--sales-backend-path", str(config.sales_backend_path),
        "--inventory-backend-path", str(config.inventory_backend_path),
        "--b2b-backend-path", str(config.b2b_backend_path),
        "--po-backend-path", str(config.po_backend_path),
        "--asin-master-path", str(config.asin_master_path),
        "--run-root", str(config.run_root),
        "--processed-dir", str(config.processed_dir),
    ])
    assert rc == 0
