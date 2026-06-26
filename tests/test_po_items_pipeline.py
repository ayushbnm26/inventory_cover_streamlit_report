from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from inventory_cover.config import PipelineConfig
from inventory_cover.exceptions import FileValidationError
from inventory_cover.io.excel_io import read_po_items_workbook
from inventory_cover.pipelines.po_items_pipeline import PoItemsPipeline


HEADERS = [
    "PO",
    "Vendor Code",
    "ASIN",
    "External ID",
    "External ID Type",
    "Ship to location",
    "Model Number",
    "Title",
    "Backordered",
    "Availability",
    "Window Type",
    "Window Start",
    "Window End",
    "Expected Date",
    "Quantity Requested",
    "Quantity Accepted",
    "Quantity Received",
    "Quantity Outstanding",
    "Unit Cost",
    "Total Cost",
]


def test_pipeline_preserves_traceability_leading_zeroes_and_duplicates(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    row = _row(
        external_id="0012345678901",
        quantity_accepted=10,
        quantity_received=4,
        quantity_outstanding=None,
    )
    _write_po_workbook(input_dir / "PurchaseOrderItems.xlsx", [row])
    _write_po_workbook(input_dir / "PurchaseOrderItems (1).xlsx", [row])

    result = PoItemsPipeline(_config(tmp_path, input_dir)).run()

    assert result.output_file.exists()
    assert result.latest_file.exists()
    assert result.backend_output_file.exists()
    assert result.backend_latest_file.exists()
    assert result.metadata_file.exists()
    assert result.log_file.exists()

    team_wb = load_workbook(result.output_file, data_only=True)
    assert team_wb.sheetnames == ["PO_Items", "Column_Guide"]
    team_headers = [cell.value for cell in team_wb["PO_Items"][1]]
    assert "Remarks" in team_headers
    assert "Run ID" not in team_headers
    assert "Source File" not in team_headers
    assert "Source Sheet" not in team_headers
    assert "Source Row" not in team_headers
    assert "Row Validation Status" not in team_headers
    assert "Row Validation Notes" not in team_headers
    team_rows = list(team_wb["PO_Items"].iter_rows(min_row=2, values_only=True))
    assert len(team_rows) == 2
    remarks_idx = team_headers.index("Remarks")
    assert team_rows[0][remarks_idx] is None
    assert team_wb["Column_Guide"].max_row > len(team_headers)
    team_wb.close()

    wb = load_workbook(result.backend_output_file, data_only=True)
    assert set(wb.sheetnames) == {
        "PO_Items_Master",
        "Run_Summary",
        "File_Audit",
        "Validation_Issues",
        "Duplicates",
    }

    master = wb["PO_Items_Master"]
    headers = [cell.value for cell in master[1]]
    assert len(headers) == len(set(headers))
    rows = list(master.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2

    header_idx = {header: idx for idx, header in enumerate(headers)}
    first = rows[0]
    assert first[header_idx["External ID"]] == "0012345678901"
    assert first[header_idx["Source File"]] == "PurchaseOrderItems.xlsx"
    assert first[header_idx["Source Row"]] == 2
    assert first[header_idx["Quantity Outstanding Normalized"]] == 0
    assert first[header_idx["Open PO Qty - Derived"]] == 6
    assert first[header_idx["Open PO Qty - Final"]] == 6
    assert first[header_idx["Row Validation Status"]] == "WARNING"
    assert "DUPLICATE_EXACT_ROW" in first[header_idx["Row Validation Notes"]]

    duplicate_rows = list(wb["Duplicates"].iter_rows(min_row=2, values_only=True))
    assert len(duplicate_rows) == 2
    assert {row[-1] for row in duplicate_rows} == {"KEPT"}

    copied_inputs = list((result.run_dir / "inputs" / "po_items").glob("*.xlsx"))
    assert len(copied_inputs) == 2
    wb.close()


def test_received_greater_than_accepted_warning(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_po_workbook(
        input_dir / "Purchase Order Items.xlsx",
        [_row(quantity_accepted=5, quantity_received=7, quantity_outstanding=0)],
    )
    _write_po_workbook(input_dir / "POItems.xlsx", [_row(po="PO-2", asin="ASIN2")])

    result = PoItemsPipeline(_config(tmp_path, input_dir)).run()

    wb = load_workbook(result.backend_output_file, data_only=True)
    issues = list(wb["Validation_Issues"].iter_rows(min_row=2, values_only=True))
    issue_types = [row[1] for row in issues]
    assert "RECEIVED_EXCEEDS_ACCEPTED" in issue_types
    wb.close()


def test_invalid_file_is_skipped_but_run_continues(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    _write_po_workbook(input_dir / "PurchaseOrderItems.xlsx", [_row()])
    _write_invalid_workbook(input_dir / "PurchaseOrderItems (1).xlsx")

    result = PoItemsPipeline(_config(tmp_path, input_dir)).run()

    wb = load_workbook(result.backend_output_file, data_only=True)
    audit_rows = list(wb["File_Audit"].iter_rows(min_row=2, values_only=True))
    statuses = {row[0]: row[8] for row in audit_rows}
    assert statuses["PurchaseOrderItems.xlsx"] == "SUCCESS"
    assert statuses["PurchaseOrderItems (1).xlsx"] == "FAILED"

    issues = list(wb["Validation_Issues"].iter_rows(min_row=2, values_only=True))
    assert any(row[1] == "FILE_SKIPPED" for row in issues)
    wb.close()


def test_missing_critical_headers_file_level_failure(tmp_path: Path) -> None:
    workbook_path = tmp_path / "bad.xlsx"
    headers = [header for header in HEADERS if header != "ASIN"]
    _write_po_workbook(workbook_path, [_row()], headers=headers)

    with pytest.raises(FileValidationError) as exc:
        read_po_items_workbook(workbook_path, _config(tmp_path, tmp_path))

    assert "ASIN" in str(exc.value)


def test_dedupe_exact_rows_drops_later_copy_with_audit(tmp_path: Path) -> None:
    input_dir = tmp_path / "incoming"
    input_dir.mkdir()
    row = _row()
    _write_po_workbook(input_dir / "PurchaseOrderItems.xlsx", [row])
    _write_po_workbook(input_dir / "PurchaseOrderItems (1).xlsx", [row])

    config = _config(tmp_path, input_dir, dedupe_exact_rows=True)
    result = PoItemsPipeline(config).run()

    wb = load_workbook(result.backend_output_file, data_only=True)
    master_rows = list(wb["PO_Items_Master"].iter_rows(min_row=2, values_only=True))
    duplicate_rows = list(wb["Duplicates"].iter_rows(min_row=2, values_only=True))

    assert len(master_rows) == 1
    assert len(duplicate_rows) == 2
    assert "DROPPED_BY_DEDUPE" in {row[-1] for row in duplicate_rows}
    wb.close()


def _config(tmp_path: Path, input_dir: Path, dedupe_exact_rows: bool = False) -> PipelineConfig:
    return PipelineConfig(
        project_root=tmp_path,
        input_dir=input_dir,
        run_root=tmp_path / "runs",
        processed_dir=tmp_path / "processed",
        dedupe_exact_rows=dedupe_exact_rows,
    )


def _row(
    po: str = "PO-1",
    asin: str = "ASIN1",
    external_id: str = "0001112223334",
    quantity_accepted: int = 10,
    quantity_received: int = 2,
    quantity_outstanding: int | None = 8,
) -> dict[str, object]:
    return {
        "PO": po,
        "Vendor Code": "VEND1",
        "ASIN": asin,
        "External ID": external_id,
        "External ID Type": "EAN",
        "Ship to location": "BLR1",
        "Model Number": "MODEL-1",
        "Title": "Synthetic Test Product",
        "Backordered": "No",
        "Availability": "Available",
        "Window Type": "Delivery",
        "Window Start": "22/6/2026",
        "Window End": "13/7/2026",
        "Expected Date": "06/08/2026",
        "Quantity Requested": 10,
        "Quantity Accepted": quantity_accepted,
        "Quantity Received": quantity_received,
        "Quantity Outstanding": quantity_outstanding,
        "Unit Cost": "INR 100.00",
        "Total Cost": "INR 1000.00",
    }


def _write_po_workbook(
    path: Path,
    rows: list[dict[str, object]],
    headers: list[str] | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PurchaseOrderItems"
    active_headers = headers or HEADERS
    ws.append(active_headers)
    for row in rows:
        ws.append([row.get(header) for header in active_headers])
    wb.save(path)


def _write_invalid_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "NotPO"
    ws.append(["hello", "world"])
    ws.append([1, 2])
    wb.save(path)
