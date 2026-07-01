from __future__ import annotations

from datetime import date
import json
from pathlib import Path
import smtplib
from typing import Any

import pytest
from openpyxl import Workbook

from inventory_cover.cli import main as cli_main
from inventory_cover.config import InventoryCoverPipelineConfig
from inventory_cover.inventory_cover_schemas import InventoryCoverPipelineRunResult, SOURCE_SUMMARY_HEADERS
from inventory_cover.io.google_drive_report_store import GoogleDriveUploadSummary
from inventory_cover.notifications import (
    EmailConfigError,
    EmailDeliveryConfig,
    EmailDeliveryError,
    EmailReportContext,
    build_email_message,
    deliver_inventory_cover_report,
)


def test_email_config_loads_from_environment() -> None:
    config = EmailDeliveryConfig.from_environment(env=_email_env(), env_file=None)

    config.validate(dry_run=False)

    assert config.smtp_host == "smtp.example.com"
    assert config.smtp_port == 587
    assert config.smtp_use_tls is True
    assert config.smtp_use_ssl is False
    assert config.to == ("ayushbnm26@gmail.com",)
    assert config.recipients == ("ayushbnm26@gmail.com",)


def test_email_config_loads_from_dotenv_file(tmp_path: Path) -> None:
    env_file = tmp_path / ".env"
    env_file.write_text(
        "\n".join(
            [
                "# local SMTP settings",
                "SMTP_HOST=smtp.example.com",
                "SMTP_PORT=465",
                "SMTP_USERNAME='<smtp-username-placeholder>'",
                "SMTP_PASSWORD=\"<smtp-password-placeholder>\"",
                "SMTP_FROM=ayush.kumar@algorithmtrix.com",
                "SMTP_USE_TLS=false",
                "SMTP_USE_SSL=true",
                "REPORT_EMAIL_TO=\"ayushbnm26@gmail.com, ops@example.com\"",
                "REPORT_EMAIL_CC=lead@example.com;finance@example.com",
                "REPORT_EMAIL_SUBJECT_PREFIX=\"Inventory Cover Report\"",
            ]
        ),
        encoding="utf-8",
    )

    config = EmailDeliveryConfig.from_environment(env={}, env_file=env_file)
    config.validate(dry_run=False)

    assert config.smtp_port == 465
    assert config.smtp_use_tls is False
    assert config.smtp_use_ssl is True
    assert config.to == ("ayushbnm26@gmail.com", "ops@example.com")
    assert config.cc == ("lead@example.com", "finance@example.com")


def test_missing_required_smtp_fields_fail_with_clear_validation_error() -> None:
    config = EmailDeliveryConfig.from_environment(env={"SMTP_PORT": "587"}, env_file=None)

    with pytest.raises(EmailConfigError) as exc:
        config.validate(dry_run=True)

    message = str(exc.value)
    assert "SMTP_HOST is required" in message
    assert "SMTP_USERNAME is required" in message
    assert "SMTP_FROM is required" in message
    assert "REPORT_EMAIL_TO is required" in message


def test_dry_run_does_not_call_smtp(tmp_path: Path) -> None:
    result = _fake_pipeline_result(tmp_path)
    config = EmailDeliveryConfig.from_environment(env=_email_env(password=""), env_file=None)

    def smtp_factory(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("SMTP should not be called during dry-run")

    delivery = deliver_inventory_cover_report(result, config, dry_run=True, smtp_factory=smtp_factory)
    audit = json.loads(delivery.audit_file.read_text(encoding="utf-8"))

    assert delivery.status == "DRY_RUN"
    assert audit["status"] == "DRY_RUN"
    assert audit["dry_run"] is True
    assert audit["attachment_path"] == str(result.team_output_file)
    assert audit["attachment_exists"] is True


def test_email_message_includes_subject_recipients_body_and_attachment(tmp_path: Path) -> None:
    attachment = tmp_path / "Inventory_Cover_Report_RUN123.xlsx"
    attachment.write_bytes(b"workbook-bytes")
    context = EmailReportContext(
        run_id="RUN123",
        generated_at="2026-06-27T10:00:00+05:30",
        product_count=2,
        validation_issue_count=1,
        warning_count=1,
        team_workbook_path=attachment,
        backend_workbook_path=tmp_path / "Inventory_Cover_Backend_Audit_RUN123.xlsx",
        report_context={
            "sales_period_start": "2026-06-01",
            "sales_period_end": "2026-06-26",
            "sales_report_updated_date": "2026-06-26",
            "inventory_period_start": "Not available",
            "inventory_period_end": "2026-06-26",
            "inventory_report_updated_date": "2026-06-26",
            "b2b_dispatch_as_of_date": "2026-06-26",
            "b2b_dispatch_lookback_start": "2026-06-25",
            "b2b_dispatch_lookback_end": "2026-06-26",
        },
        missing_fields=("inventory_period_start",),
    )
    config = EmailDeliveryConfig.from_environment(env=_email_env(cc="cc@example.com"), env_file=None)

    message = build_email_message(context, config, mail_timestamp="2026-06-27T10:01:00+05:30")

    assert message["Subject"] == "Inventory Cover Report - 2026-06-27"
    assert message["To"] == "ayushbnm26@gmail.com"
    assert message["Cc"] == "cc@example.com"
    body = message.get_body(preferencelist=("plain",)).get_content()
    assert "Hi," in body
    assert "Please find attached the Inventory Cover Report for 2026-06-27." in body
    assert "- Sales period: 2026-06-01 to 2026-06-26" in body
    assert "- Inventory report updated: 2026-06-26" in body
    attachments = list(message.iter_attachments())
    assert len(attachments) == 1
    assert attachments[0].get_filename() == "Inventory_Cover_Report_RUN123.xlsx"


def test_email_body_excludes_dashboard_url_when_unset(tmp_path: Path) -> None:
    attachment = tmp_path / "Inventory_Cover_Report_RUN123.xlsx"
    attachment.write_bytes(b"workbook-bytes")
    context = _email_context(attachment, tmp_path)
    config = EmailDeliveryConfig.from_environment(env=_email_env(), env_file=None)

    message = build_email_message(context, config, mail_timestamp="2026-06-27T10:01:00+05:30")
    body = message.get_body(preferencelist=("plain",)).get_content()

    assert "Dashboard view:" not in body
    assert "read-only visual summary" not in body


def test_email_body_includes_dashboard_url_when_set(tmp_path: Path) -> None:
    attachment = tmp_path / "Inventory_Cover_Report_RUN123.xlsx"
    attachment.write_bytes(b"workbook-bytes")
    context = _email_context(attachment, tmp_path)
    config = EmailDeliveryConfig.from_environment(
        env=_email_env(dashboard_url="https://inventory-cover.streamlit.app"),
        env_file=None,
    )

    message = build_email_message(context, config, mail_timestamp="2026-06-27T10:01:00+05:30")
    body = message.get_body(preferencelist=("plain",)).get_content()

    assert "Dashboard view: https://inventory-cover.streamlit.app" in body
    assert "attached Excel workbook remains the official working report" in body


def test_attachment_must_exist_before_sending(tmp_path: Path) -> None:
    result = _fake_pipeline_result(tmp_path, attachment_exists=False)
    config = EmailDeliveryConfig.from_environment(env=_email_env(password=""), env_file=None)

    with pytest.raises(EmailDeliveryError) as exc:
        deliver_inventory_cover_report(result, config, dry_run=True)

    assert "Attachment does not exist" in str(exc.value)
    audit = json.loads((result.run_dir / "notifications" / "email_delivery.json").read_text(encoding="utf-8"))
    assert audit["status"] == "FAILED"
    assert audit["attachment_exists"] is False


def test_cli_without_email_flag_creates_no_notification_artifacts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setenv("GDRIVE_ENABLED", "false")
    config = _make_inventory_cover_sources(tmp_path)

    rc = cli_main(_inventory_cover_cli_args(config))

    assert rc == 0
    run_dirs = list((tmp_path / "runs").iterdir())
    assert len(run_dirs) == 1
    assert not (run_dirs[0] / "notifications").exists()


def test_cli_send_email_dry_run_writes_audit_metadata(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("GDRIVE_ENABLED", "false")
    config = _make_inventory_cover_sources(tmp_path)
    for key, value in _email_env(password="").items():
        monkeypatch.setenv(key, value)

    rc = cli_main(_inventory_cover_cli_args(config) + ["--send-email", "--email-dry-run"])

    assert rc == 0
    run_dir = next((tmp_path / "runs").iterdir())
    audit = json.loads((run_dir / "notifications" / "email_delivery.json").read_text(encoding="utf-8"))
    assert audit["status"] == "DRY_RUN"
    assert audit["dry_run"] is True
    assert audit["attachment_path"].endswith(f"Inventory_Cover_Report_{audit['run_id']}.xlsx")
    assert audit["report_context"]["sales_period_start"] == "2024-05-01"
    assert audit["report_context"]["sales_report_updated_date"] == "2024-05-31"
    assert "SMTP_PASSWORD" not in json.dumps(audit)


def test_cli_drive_upload_is_config_gated(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    config = _make_inventory_cover_sources(tmp_path)
    calls: dict[str, Any] = {}

    def fake_upload(result: InventoryCoverPipelineRunResult, drive_config: Any) -> GoogleDriveUploadSummary:
        calls["enabled"] = drive_config.enabled
        calls["folder_id"] = drive_config.folder_id
        calls["team_latest_exists"] = result.team_latest_file.exists()
        return GoogleDriveUploadSummary(
            run_id=result.run_id,
            enabled=True,
            status="SUCCESS",
            audit_file=result.run_dir / "notifications" / "google_drive_upload.json",
            log_file=result.run_dir / "logs" / "google_drive_upload.log",
            started_at="2026-07-01T10:00:00+05:30",
            completed_at="2026-07-01T10:00:01+05:30",
            duration_seconds=1.0,
        )

    monkeypatch.setenv("GDRIVE_ENABLED", "true")
    monkeypatch.setenv("GDRIVE_FOLDER_ID", "folder123")
    monkeypatch.setattr("inventory_cover.cli.upload_inventory_cover_reports_to_drive", fake_upload)

    rc = cli_main(_inventory_cover_cli_args(config))

    assert rc == 0
    assert calls == {"enabled": True, "folder_id": "folder123", "team_latest_exists": True}


def test_smtp_failure_writes_failure_audit_without_deleting_workbook(tmp_path: Path) -> None:
    result = _fake_pipeline_result(tmp_path)
    username_marker = "USERNAME_MARKER_FOR_REDACTION"
    password_marker = "PASSWORD_MARKER_FOR_REDACTION"
    config = EmailDeliveryConfig.from_environment(
        env=_email_env(username=username_marker, password=password_marker),
        env_file=None,
    )

    with pytest.raises(EmailDeliveryError):
        deliver_inventory_cover_report(result, config, dry_run=False, smtp_factory=FailingSMTP)

    assert result.team_output_file.exists()
    audit_path = result.run_dir / "notifications" / "email_delivery.json"
    audit_text = audit_path.read_text(encoding="utf-8")
    audit = json.loads(audit_text)
    log_text = (result.run_dir / "logs" / "email_delivery.log").read_text(encoding="utf-8")
    assert audit["status"] == "FAILED"
    assert audit["error_type"] == "SMTP_AUTHENTICATION_FAILED"
    assert password_marker not in audit_text
    assert username_marker not in audit_text
    assert password_marker not in log_text
    assert username_marker not in log_text


class FailingSMTP:
    def __init__(self, host: str, port: int, timeout: float) -> None:
        self.host = host
        self.port = port
        self.timeout = timeout

    def __enter__(self) -> "FailingSMTP":
        return self

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        return None

    def starttls(self) -> None:
        return None

    def login(self, username: str, password: str) -> None:
        raise smtplib.SMTPAuthenticationError(
            535,
            f"auth failed for {username} using {password}".encode(),
        )

    def send_message(self, *args: Any, **kwargs: Any) -> None:
        raise AssertionError("send_message should not run after login failure")


def _email_env(
    *,
    username: str = "<smtp-username-placeholder>",
    password: str = "<smtp-password-placeholder>",
    cc: str = "",
    dashboard_url: str = "",
) -> dict[str, str]:
    return {
        "SMTP_HOST": "smtp.example.com",
        "SMTP_PORT": "587",
        "SMTP_USERNAME": username,
        "SMTP_PASSWORD": password,
        "SMTP_FROM": "ayush.kumar@algorithmtrix.com",
        "SMTP_USE_TLS": "true",
        "SMTP_USE_SSL": "false",
        "SMTP_TIMEOUT_SECONDS": "30",
        "REPORT_EMAIL_TO": "ayushbnm26@gmail.com",
        "REPORT_EMAIL_CC": cc,
        "REPORT_EMAIL_BCC": "",
        "REPORT_EMAIL_REPLY_TO": "",
        "REPORT_EMAIL_SUBJECT_PREFIX": "Inventory Cover Report",
        "INVENTORY_DASHBOARD_URL": dashboard_url,
    }


def _email_context(attachment: Path, tmp_path: Path) -> EmailReportContext:
    return EmailReportContext(
        run_id="RUN123",
        generated_at="2026-06-27T10:00:00+05:30",
        product_count=2,
        validation_issue_count=1,
        warning_count=1,
        team_workbook_path=attachment,
        backend_workbook_path=tmp_path / "Inventory_Cover_Backend_Audit_RUN123.xlsx",
        report_context={
            "sales_period_start": "2026-06-01",
            "sales_period_end": "2026-06-26",
            "sales_report_updated_date": "2026-06-26",
            "inventory_period_start": "Not available",
            "inventory_period_end": "2026-06-26",
            "inventory_report_updated_date": "2026-06-26",
            "b2b_dispatch_as_of_date": "2026-06-26",
            "b2b_dispatch_lookback_start": "2026-06-25",
            "b2b_dispatch_lookback_end": "2026-06-26",
        },
        missing_fields=("inventory_period_start",),
    )


def _fake_pipeline_result(tmp_path: Path, *, attachment_exists: bool = True) -> InventoryCoverPipelineRunResult:
    run_id = "RUN123"
    run_dir = tmp_path / "runs" / run_id
    outputs = run_dir / "outputs" / "inventory_cover"
    metadata_dir = run_dir / "metadata"
    validation_dir = run_dir / "validation"
    outputs.mkdir(parents=True)
    metadata_dir.mkdir(parents=True)
    validation_dir.mkdir(parents=True)
    team_output = outputs / f"Inventory_Cover_Report_{run_id}.xlsx"
    if attachment_exists:
        team_output.write_bytes(b"workbook-bytes")
    backend_output = outputs / f"Inventory_Cover_Backend_Audit_{run_id}.xlsx"
    b2b_backend = run_dir / "inputs" / "inventory_cover" / "B2B_Dispatch_Backend_Audit_latest.xlsx"
    _write_b2b_backend_with_summary(b2b_backend)
    _write_inventory_cover_backend_summary(backend_output, b2b_backend)
    metadata_file = metadata_dir / "run_metadata.json"
    metadata_file.write_text(json.dumps({"end_time": "2026-06-27T10:00:00"}), encoding="utf-8")
    validation_file = validation_dir / "inventory_cover_validation_issues.json"
    validation_file.write_text(json.dumps({"issues": []}), encoding="utf-8")

    return InventoryCoverPipelineRunResult(
        run_id=run_id,
        run_dir=run_dir,
        team_output_file=team_output,
        team_latest_file=tmp_path / "processed" / "latest" / "Inventory_Cover_Report_latest.xlsx",
        backend_output_file=backend_output,
        backend_latest_file=tmp_path / "processed" / "latest" / "Inventory_Cover_Backend_Audit_latest.xlsx",
        metadata_file=metadata_file,
        validation_file=validation_file,
        log_file=run_dir / "logs" / "inventory_cover_pipeline.log",
        product_count=2,
        validation_issue_count=1,
        warning_count=1,
    )


def _write_inventory_cover_backend_summary(path: Path, copied_b2b_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Source_Summary"
    ws.append(list(SOURCE_SUMMARY_HEADERS))
    ws.append(
        [
            "RUN123",
            "2026-06-27T10:00:00",
            "Sales",
            "sales.xlsx",
            "sales.xlsx",
            "Yes",
            "Sales_Master",
            1,
            1,
            1,
            date(2026, 6, 1),
            date(2026, 6, 26),
            date(2026, 6, 26),
            "FRESH",
            "",
        ]
    )
    ws.append(
        [
            "RUN123",
            "2026-06-27T10:00:00",
            "Inventory",
            "inventory.xlsx",
            "inventory.xlsx",
            "Yes",
            "Inventory_Master",
            1,
            1,
            1,
            None,
            date(2026, 6, 26),
            date(2026, 6, 26),
            "FRESH",
            "",
        ]
    )
    ws.append(
        [
            "RUN123",
            "2026-06-27T10:00:00",
            "B2B Dispatch",
            str(copied_b2b_path),
            str(copied_b2b_path),
            "Yes",
            "B2B_Dispatch_Master",
            1,
            1,
            1,
            None,
            None,
            None,
            "FRESH",
            "",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_b2b_backend_with_summary(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "B2B_Dispatch_Master"
    ws.append(["ASIN", "Dispatch Qty", "Dispatch Date", "Included In Lookback Window"])
    ws.append(["A1", 5, date(2026, 6, 26), True])
    summary = wb.create_sheet("Run_Summary")
    summary.append(["Run ID", "As of date", "Lookback start date", "Lookback end date"])
    summary.append(["B2B123", date(2026, 6, 26), date(2026, 6, 25), date(2026, 6, 26)])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _make_inventory_cover_sources(tmp_path: Path) -> InventoryCoverPipelineConfig:
    data_dir = tmp_path / "data"
    sales_path = data_dir / "sales.xlsx"
    inventory_path = data_dir / "inventory.xlsx"
    b2b_path = data_dir / "b2b.xlsx"
    po_path = data_dir / "po.xlsx"
    asin_path = data_dir / "asin.xlsx"
    _write_workbook(
        sales_path,
        "Sales_Master",
        [
            "Run ID",
            "Source File",
            "Source Sheet",
            "Source Row",
            "ASIN",
            "Model Number",
            "Viewing Range Start",
            "Viewing Range End",
            "Report Updated Date",
            "Shipped Units",
        ],
        [
            {
                "ASIN": "A1",
                "Model Number": "M-A1",
                "Viewing Range Start": date(2024, 5, 1),
                "Viewing Range End": date(2024, 5, 30),
                "Report Updated Date": date(2024, 5, 31),
                "Shipped Units": 30,
            }
        ],
    )
    _write_workbook(
        inventory_path,
        "Inventory_Master",
        [
            "Run ID",
            "Source File",
            "Source Sheet",
            "Source Row",
            "ASIN",
            "Model Number",
            "Viewing Range End",
            "Report Updated Date",
            "Sellable On Hand Units",
            "Sellable In Transit Units",
        ],
        [
            {
                "ASIN": "A1",
                "Model Number": "M-A1",
                "Viewing Range End": date(2024, 5, 31),
                "Report Updated Date": date(2024, 5, 31),
                "Sellable On Hand Units": 100,
                "Sellable In Transit Units": 10,
            }
        ],
    )
    _write_workbook(
        b2b_path,
        "B2B_Dispatch_Master",
        ["Run ID", "Source File", "Source Sheet", "Source Row", "ASIN", "Model Number", "Dispatch Qty", "Included In Lookback Window"],
        [{"ASIN": "A1", "Model Number": "M-A1", "Dispatch Qty": 5, "Included In Lookback Window": True}],
    )
    _write_workbook(
        po_path,
        "PO_Items_Master",
        ["Run ID", "Source File", "Source Sheet", "Source Row", "ASIN", "Model Number", "Open PO Qty - Final"],
        [{"ASIN": "A1", "Model Number": "M-A1", "Open PO Qty - Final": 20}],
    )
    _write_workbook(
        asin_path,
        "ASIN_Master",
        ["ASIN", "SKU", "Brand", "Brand Name", "Vendor", "Main Category", "Sub Category", "Aligned DOH Target"],
        [{"ASIN": "A1", "SKU": "SKU-A1", "Brand": "BR", "Brand Name": "Brand One", "Aligned DOH Target": 45}],
    )
    return InventoryCoverPipelineConfig(
        project_root=tmp_path,
        sales_backend_path=sales_path,
        inventory_backend_path=inventory_path,
        b2b_backend_path=b2b_path,
        po_backend_path=po_path,
        asin_master_path=asin_path,
        run_root=tmp_path / "runs",
        processed_dir=tmp_path / "processed" / "inventory_cover",
    )


def _write_workbook(path: Path, sheet_name: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _inventory_cover_cli_args(config: InventoryCoverPipelineConfig) -> list[str]:
    return [
        "run-inventory-cover",
        "--sales-backend-path",
        str(config.sales_backend_path),
        "--inventory-backend-path",
        str(config.inventory_backend_path),
        "--b2b-backend-path",
        str(config.b2b_backend_path),
        "--po-backend-path",
        str(config.po_backend_path),
        "--asin-master-path",
        str(config.asin_master_path),
        "--run-root",
        str(config.run_root),
        "--processed-dir",
        str(config.processed_dir),
    ]
